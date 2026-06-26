from django.shortcuts import render
from rest_framework.response import Response
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import FileResponse, HttpResponse
from rest_framework.decorators import api_view, parser_classes
from rest_framework.response import Response
from .models import Campaign, CampaignExcel
from django.core.files.base import ContentFile
from datetime import datetime
from campaigns.models import LineItem

# Create your views here.


def parse_date(date_str):

    if not date_str:
        return None
    return datetime.fromisoformat(date_str.replace("Z", "")).date()


def build_campaign_excel(campaign, report_type="cpm", overrides=None) -> bytes:
    """
    overrides = { "LIBILL003": { "impressions": 50000, "start_date": "2026-06-01", ... } }
    """
    overrides = overrides or {}
    wb = Workbook()
    wb.remove(wb.active)

    # ── Styles (same as before) ──
    header_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    label_fill = PatternFill("solid", start_color="D9E1F2")
    title_fill = PatternFill("solid", start_color="1F4E79")
    title_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    border_side = Side(style="thin", color="B8CCE4")
    thin_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    line_items = campaign.line_items.all()

    for li in line_items:
        sheet_name = str(li.line_item_id)[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40

        ws.merge_cells("A1:B1")
        ws["A1"] = f"Campaign Report — {campaign.campaign_id}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 22

        insertion_order = getattr(campaign, "insertion_order", None)
        is_cpc = report_type == "cpc"
        has_creatives = (
            li.creatives_detail.exists() or li.third_party_creatives.exists()
        )
        creative_msg = (
            "Creatives will be shared by Creatives Team"
            if has_creatives
            else "Not yet received"
        )

        # ✅ Get overrides for this line item
        li_overrides = overrides.get(li.line_item_id, {})

        # ✅ Use override value if present, else use DB value
        def val(field, db_value):
            return li_overrides.get(field, db_value)

        impressions = val("impressions", li.impressions)
        start_date = val("start_date", str(li.start_date) if li.start_date else "")
        end_date = val("end_date", str(li.end_date) if li.end_date else "")
        units = val("units", li.units or "")
        ctr_raw = val("ctr", li.ctr)
        viewability_raw = val("viewability", li.viewability)
        vcr_raw = val("vcr", li.vcr)
        kpi_notes = val("kpi_notes", li.kpi_notes or "")
        sitelist = val("sitelist", "")

        ctr_display = f"{ctr_raw}%" if ctr_raw else ""
        viewability_display = f"{viewability_raw}%" if viewability_raw else ""
        vcr_display = f"{vcr_raw}%" if vcr_raw else ""

        if is_cpc:
            rows = [
                (
                    "Date of ID Setup",
                    (
                        campaign.created_at.strftime("%d-%B-%Y")
                        if campaign.created_at
                        else ""
                    ),
                ),
                ("Advertiser ID", campaign.client_campaign_ID or ""),
                ("Campaign ID", f"{campaign.campaign_id}-A"),
                ("IO ID", insertion_order.io_id if insertion_order else ""),
                ("IO Name", campaign.campaign_name or ""),
                ("Clicks Booked", impressions or ""),
                ("Start Date", start_date),
                ("End Date", end_date),
                ("Target CPC", ctr_display),
                ("Booked Budget", ""),  # ← CPC: Booked Budget row, empty value
                ("Line Item ID", li.line_item_id or ""),
                ("Line Item Name", li.line_item_name or ""),
                ("Ethnicity", li.ethnicity or ""),
                ("Ad Format", li.ad_format or ""),
                ("Geography", _parse_geo(li.geo_targeting) if li.geo_targeting else ""),
                ("Market", ""),
                ("Viewability", viewability_display),
                ("Creative", creative_msg),
                ("VCR", vcr_display),
                ("Daily Clicks", ""),  # ← CPC only row
                ("Sitelist", sitelist),
            ]
        else:
            rows = [
                (
                    "Date of ID Setup",
                    (
                        campaign.created_at.strftime("%d-%B-%Y")
                        if campaign.created_at
                        else ""
                    ),
                ),
                ("Advertiser ID", campaign.client_campaign_ID or ""),
                ("Campaign ID", campaign.campaign_id),
                ("IO ID", insertion_order.io_id if insertion_order else ""),
                ("IO Name", campaign.campaign_name or ""),
                ("Impressions Booked", impressions or ""),
                ("Start Date", start_date),
                ("End Date", end_date),
                ("Target CPM", ctr_display),
                ("Target CTR", ctr_display),  # ← CPM: Target CTR row
                ("Line Item ID", li.line_item_id or ""),
                ("Line Item Name", li.line_item_name or ""),
                ("Ethnicity", li.ethnicity or ""),
                ("Ad Format", li.ad_format or ""),
                ("Geography", _parse_geo(li.geo_targeting) if li.geo_targeting else ""),
                ("Market", ""),
                ("Viewability", viewability_display),
                ("Creative", creative_msg),
                ("VCR", vcr_display),
                ("KPI", kpi_notes),
                ("Sitelist", sitelist),
            ]
        for i, (label, value) in enumerate(rows, start=2):
            label_cell = ws.cell(row=i, column=1, value=label)
            value_cell = ws.cell(row=i, column=2, value=value)
            label_cell.font = header_font
            label_cell.fill = label_fill
            label_cell.alignment = left_align
            label_cell.border = thin_border
            value_cell.font = value_font
            value_cell.alignment = left_align
            value_cell.border = thin_border
            ws.row_dimensions[i].height = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _parse_geo(geo_targeting):
    """Parse geo_targeting field (JSON string or list) into readable string."""
    try:
        if isinstance(geo_targeting, str):
            data = json.loads(geo_targeting)
        else:
            data = geo_targeting
        parts = []
        for loc in (data if isinstance(data, list) else [data]):
            segment = ", ".join(
                filter(
                    None,
                    [loc.get("country", ""), loc.get("state", ""), loc.get("city", "")],
                )
            )
            if segment:
                parts.append(segment)
        return " | ".join(parts) if parts else ""
    except Exception:
        return str(geo_targeting) if geo_targeting else ""


# ── Generate & store Excel ──────────────────────────────────────────────────


@api_view(["POST"])
def generate_campaign_excel(request, campaign_id):
    try:
        campaign = (
            Campaign.objects.select_related("client", "insertion_order")
            .prefetch_related(
                "line_items",
                "line_items__creatives_detail",
                "line_items__third_party_creatives",
            )
            .get(campaign_id=campaign_id)
        )
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    report_type = request.data.get("report_type", "cpm")

    # ── Get existing overrides if record exists ──
    existing_overrides = {}
    try:
        existing_excel = CampaignExcel.objects.get(
            campaign=campaign, report_type=report_type
        )
        existing_overrides = existing_excel.line_item_overrides or {}
    except CampaignExcel.DoesNotExist:
        pass

    # ── Always regenerate with existing overrides ──
    excel_bytes = build_campaign_excel(
        campaign, report_type, overrides=existing_overrides
    )
    filename = f"{campaign_id}_{report_type}.xlsx"

    # ── get_or_create then always save the file ──
    excel_obj, _ = CampaignExcel.objects.get_or_create(
        campaign=campaign, report_type=report_type, defaults={"line_item_overrides": {}}
    )
    excel_obj.excel_file.save(filename, ContentFile(excel_bytes), save=True)

    url = request.build_absolute_uri(excel_obj.excel_file.url)
    return Response(
        {
            "message": "Excel generated successfully",
            "campaign_id": campaign_id,
            "download_url": url,
            "filename": filename,
            "generated_at": excel_obj.generated_at.isoformat(),
        },
        status=200,
    )


@api_view(["GET"])
def get_campaigns_excel_list(request):
    campaigns = (
        Campaign.objects.select_related("client")
        .filter(approval_status="approved")
        .prefetch_related("excel_reports", "line_items")
        .order_by("-created_at")
    )

    data = []
    for c in campaigns:
        excel_map = {e.report_type: e for e in c.excel_reports.all()}

        # ✅ Check units across all line items
        all_units = [(li.units or "").upper() for li in c.line_items.all() if li.units]

        has_cpm = any(u == "CPM" for u in all_units)
        has_cpc = any(u == "CPC" for u in all_units)

        # ✅ Decide which report types to generate
        if has_cpm and has_cpc:
            report_types = ["cpm", "cpc"]  # both units → both reports
        elif has_cpc:
            report_types = ["cpc"]  # only CPC → one report (no -A suffix)
        else:
            report_types = ["cpm", "cpc"]  # only CPM or unknown → both reports

        for report_type in report_types:
            excel = excel_map.get(report_type)
            suffix = "-A" if report_type == "cpc" else ""
            data.append(
                {
                    "campaign_id": f"{c.campaign_id}{suffix}",
                    "campaign_id_raw": c.campaign_id,
                    "report_type": report_type,
                    "campaign_name": c.campaign_name,
                    "client_name": c.client.name if c.client else "",
                    "client_id": c.client.client_id if c.client else "",
                    "start_date": str(c.start_date) if c.start_date else "",
                    "end_date": str(c.end_date) if c.end_date else "",
                    "line_items_count": c.line_items.count(),
                    "excel_generated": excel is not None,
                    "excel_url": (
                        request.build_absolute_uri(excel.excel_file.url)
                        if excel
                        else None
                    ),
                    "generated_at": excel.generated_at.isoformat() if excel else None,
                    "publish_status": excel.publish_status if excel else None,
                }
            )

    return Response(data)


# ── Download saved Excel ────────────────────────────────────────────────────


@api_view(["GET"])
def download_campaign_excel(request, campaign_id):
    # ✅ Read report_type from query param
    report_type = request.query_params.get("report_type", "cpm")

    try:
        excel_obj = CampaignExcel.objects.get(
            campaign__campaign_id=campaign_id,
            report_type=report_type,  # ✅ Add this filter
        )
    except CampaignExcel.DoesNotExist:
        # Generate on-the-fly if not saved yet
        try:
            campaign = (
                Campaign.objects.select_related("client", "insertion_order")
                .prefetch_related("line_items")
                .get(campaign_id=campaign_id)
            )
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)

        excel_bytes = build_campaign_excel(campaign, report_type)
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{campaign_id}_{report_type}.xlsx"'
        )
        return response

    response = FileResponse(
        excel_obj.excel_file.open("rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{campaign_id}_{report_type}.xlsx"'
    )
    return response


@api_view(["POST"])
def save_excel_edits_to_db(request, campaign_id):
    try:
        report_type = request.data.get("report_type", "cpm")
        line_item_id = request.data.get("line_item_id")

        if not line_item_id:
            return Response(
                {"error": "line_item_id is required"},
                status=400,
            )

        # ── 1. Get CampaignExcel record ──
        try:
            excel_obj = CampaignExcel.objects.get(
                campaign__campaign_id=campaign_id,
                report_type=report_type,
            )
        except CampaignExcel.DoesNotExist:
            return Response(
                {"error": "Excel record not found. Generate first."},
                status=404,
            )

        # ── 2. Get LineItem (read-only, never modified) ──
        try:
            line_item = LineItem.objects.get(line_item_id=line_item_id)
        except LineItem.DoesNotExist:
            return Response(
                {"error": "LineItem not found"},
                status=404,
            )

        # ── 3. Get Campaign ──
        try:
            campaign = (
                Campaign.objects.select_related(
                    "client",
                    "insertion_order",
                )
                .prefetch_related(
                    "line_items",
                    "line_items__creatives_detail",
                    "line_items__third_party_creatives",
                )
                .get(campaign_id=campaign_id)
            )
        except Campaign.DoesNotExist:
            return Response(
                {"error": "Campaign not found"},
                status=404,
            )

        editable_fields = [
            "impressions",
            "start_date",
            "end_date",
            "units",
            "ctr",
            "viewability",
            "vcr",
            "kpi_notes",
            "sitelist",
            "advertiser_id",
            "target_cpm",
            "target_ctr",
            "target_cpc",
            "booked_budget",
        ]

        # ── 4. Build override dict ──
        li_override = excel_obj.line_item_overrides.get(
            line_item_id,
            {},
        )

        for field in editable_fields:
            if field in request.data and request.data[field] is not None:
                li_override[field] = request.data[field]

        # ── 5. Save overrides to CampaignExcel.line_item_overrides ──
        overrides = excel_obj.line_item_overrides.copy()
        overrides[line_item_id] = li_override

        excel_obj.line_item_overrides = overrides
        excel_obj.save(update_fields=["line_item_overrides"])

        # ── 6. Save ONLY to CampaignLineItemExcelDetails ──
        from .models import CampaignLineItemExcelDetails

        def safe_int(value):
            try:
                return int(str(value).replace(",", "").strip())
            except Exception:
                return None

        def safe_float(value):
            try:
                return float(str(value).replace("%", "").strip())
            except Exception:
                return None

        excel_li_obj, _ = CampaignLineItemExcelDetails.objects.update_or_create(
            line_item=line_item,
            report_type=report_type,
            defaults={
                "campaign": campaign,
                # CPM report → impressions
                "impressions": (
                    safe_int(li_override["impressions"])
                    if report_type == "cpm" and "impressions" in li_override
                    else None
                ),
                # CPC report → clicks
                "clicks": (
                    safe_int(li_override["impressions"])
                    if report_type == "cpc" and "impressions" in li_override
                    else None
                ),
                "start_date": (
                    parse_date(str(li_override["start_date"]))
                    if li_override.get("start_date")
                    else None
                ),
                "end_date": (
                    parse_date(str(li_override["end_date"]))
                    if li_override.get("end_date")
                    else None
                ),
                "advertiser_id": (
                    li_override.get("advertiser_id")
                    or campaign.client_campaign_ID
                    or ""
                ),
                "target_cpm": (
                    safe_float(li_override["target_cpm"])
                    if "target_cpm" in li_override
                    else None
                ),
                "target_ctr": (
                    safe_float(li_override["target_ctr"])
                    if "target_ctr" in li_override
                    else None
                ),
                "target_cpc": (
                    safe_float(li_override["target_cpc"])
                    if report_type == "cpc" and "target_cpc" in li_override
                    else None
                ),
                "booked_budget": (
                    safe_float(li_override["booked_budget"])
                    if report_type == "cpc" and "booked_budget" in li_override
                    else None
                ),
                "sitelist": li_override.get(
                    "sitelist",
                    "",
                ),
            },
        )

        # ── 7. Re-generate Excel with updated overrides ──
        excel_bytes = build_campaign_excel(
            campaign,
            report_type,
            overrides=overrides,
        )

        filename = f"{campaign_id}_{report_type}.xlsx"

        excel_obj.excel_file.save(
            filename,
            ContentFile(excel_bytes),
            save=True,
        )

        url = request.build_absolute_uri(excel_obj.excel_file.url)

        return Response(
            {
                "message": "Saved to CampaignLineItemExcelDetails only (LineItem unchanged)",
                "download_url": url,
                "line_item_overrides": overrides,
                "excel_line_item_id": excel_li_obj.id,
            },
            status=200,
        )

    except Exception as e:
        import traceback

        print(traceback.format_exc())

        return Response(
            {"error": str(e)},
            status=500,
        )


@api_view(["GET"])
def get_line_item_excel_data(request, campaign_id):
    from .models import CampaignLineItemExcelDetails

    report_type = request.query_params.get("report_type", "cpm")

    try:
        campaign = (
            Campaign.objects.select_related("client")
            .prefetch_related("line_items")
            .get(campaign_id=campaign_id)
        )
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    data = {}
    for li in campaign.line_items.all():
        try:
            # ── Already saved → return saved values ──
            excel_data = CampaignLineItemExcelDetails.objects.get(
                line_item=li, report_type=report_type
            )
            data[li.line_item_id] = {
                "impressions": excel_data.impressions,
                "clicks": excel_data.clicks,
                "start_date": (
                    str(excel_data.start_date) if excel_data.start_date else ""
                ),
                "end_date": str(excel_data.end_date) if excel_data.end_date else "",
                "advertiser_id": excel_data.advertiser_id
                or campaign.client_campaign_ID
                or "",
                "target_cpm": excel_data.target_cpm,
                "target_ctr": excel_data.target_ctr,
                "target_cpc": excel_data.target_cpc,
                "booked_budget": excel_data.booked_budget,
                "sitelist": excel_data.sitelist or "",
            }
        except CampaignLineItemExcelDetails.DoesNotExist:
            # ── Not saved yet → return defaults from Campaign + LineItem ──
            data[li.line_item_id] = {
                "impressions": li.impressions,
                "clicks": None,
                "start_date": str(li.start_date) if li.start_date else "",
                "end_date": str(li.end_date) if li.end_date else "",
                "advertiser_id": campaign.client_campaign_ID or "",
                "target_cpm": li.unit_cost or "",  # unit_cost → target_cpm
                "target_ctr": li.ctr or "",
                "target_cpc": None,
                "booked_budget": None,
                "sitelist": "",
            }

    return Response(data)


@api_view(["POST"])
def publish_campaign_excel(request, campaign_id):
    report_type = request.data.get("report_type", "cpm")
    try:
        excel_obj = CampaignExcel.objects.get(
            campaign__campaign_id=campaign_id, report_type=report_type
        )
    except CampaignExcel.DoesNotExist:
        return Response({"error": "Excel not found. Generate first."}, status=404)

    excel_obj.publish_status = "published"
    excel_obj.save(update_fields=["publish_status"])

    return Response(
        {
            "message": "Excel published successfully",
            "campaign_id": campaign_id,
            "report_type": report_type,
            "publish_status": excel_obj.publish_status,
        },
        status=200,
    )


from datetime import timedelta
from .models import CampaignDailyEntry, CampaignDailyReportExcel


def daterange(start_date, end_date):
    days = (end_date - start_date).days
    for i in range(days + 1):
        yield start_date + timedelta(days=i)


# ── NEW: Build the Daily Reports Excel — one sheet per campaign (sheet name = io_id),
# all line items laid out as side-by-side column blocks with a daily date table ──
import re


def _safe_sheet_name(name: str, used_names: set) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]. Dedupe within a workbook."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", str(name)).strip() or "Sheet"
    base = cleaned[:31]
    candidate = base
    suffix = 1
    while candidate in used_names:
        suffix_str = f"~{suffix}"
        candidate = base[: 31 - len(suffix_str)] + suffix_str
        suffix += 1
    used_names.add(candidate)
    return candidate


def _is_video_line_item(li) -> bool:
    """Detect video line items from the raw ad_format JSONField (list or string)."""
    raw = li.ad_format
    if isinstance(raw, list):
        text = " ".join(str(x) for x in raw)
    else:
        text = str(raw or "")
    text = text.lower()
    return "video" in text or "youtube" in text


def build_daily_report_excel(campaign) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    label_fill = PatternFill("solid", start_color="D9E1F2")
    border_side = Side(style="thin", color="B8CCE4")
    thin_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    table_header_fill = PatternFill("solid", start_color="1F4E79")
    table_header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    insertion_order = getattr(campaign, "insertion_order", None)
    main_sheet_name = (
        str(insertion_order.io_id)[:31]
        if insertion_order
        else str(campaign.campaign_id)[:31]
    )

    used_sheet_names = {main_sheet_name}
    ws_main = wb.create_sheet(title=main_sheet_name)

    line_items = list(campaign.line_items.all())

    entries = CampaignDailyEntry.objects.filter(campaign=campaign).select_related(
        "line_item"
    )
    entries_map = {}
    for e in entries:
        entries_map.setdefault(e.line_item_id, {})[e.entry_date] = e

    # ── Split line items: video ones get their own sheet, the rest share the main sheet ──
    video_line_items = [li for li in line_items if _is_video_line_item(li)]
    non_video_line_items = [li for li in line_items if not _is_video_line_item(li)]

    def write_summary_header(
        ws,
        start_col,
        li,
        target_impressions,
        first_data_row,
        last_data_row,
        col_letters,
        daily_target_col_idx,
    ):
        """Rows 3-8: shared summary block layout, reused by both narrow (banner) and wide (video) sheets."""
        col_a = col_letters[0]
        col_d = col_letters[
            daily_target_col_idx
        ]  # the "value" column for daily report summary numbers

        ws.cell(row=3, column=start_col, value=li.start_date.strftime("%d %B, %Y"))
        ws.merge_cells(
            start_row=3, start_column=start_col + 1, end_row=3, end_column=start_col + 2
        )
        ws.cell(row=3, column=start_col + 1, value=li.line_item_name or li.line_item_id)
        ws.cell(row=3, column=start_col + 1).alignment = center_align
        ws.cell(row=3, column=start_col + daily_target_col_idx, value="Daily")
        for c in range(start_col, start_col + daily_target_col_idx + 1):
            cell = ws.cell(row=3, column=c)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.cell(row=4, column=start_col, value=li.end_date.strftime("%d %B, %Y"))
        ws.merge_cells(
            start_row=4, start_column=start_col + 1, end_row=4, end_column=start_col + 2
        )
        ws.cell(row=4, column=start_col + 1, value="Target impressions")
        ws.cell(
            row=4, column=start_col + daily_target_col_idx, value=target_impressions
        )

        ws.merge_cells(
            start_row=5, start_column=start_col + 1, end_row=5, end_column=start_col + 2
        )
        ws.cell(row=5, column=start_col + 1, value="Achieved impressions")
        impressions_col_letter = col_letters[1]  # the Impressions data column
        ws.cell(
            row=5,
            column=start_col + daily_target_col_idx,
            value=f"=SUM({impressions_col_letter}{first_data_row}:{impressions_col_letter}{last_data_row})",
        )

        ws.merge_cells(
            start_row=6, start_column=start_col + 1, end_row=6, end_column=start_col + 2
        )
        ws.cell(row=6, column=start_col + 1, value="Remaining impressions")
        ws.cell(
            row=6, column=start_col + daily_target_col_idx, value=f"={col_d}4-{col_d}5"
        )

        ws.merge_cells(
            start_row=7, start_column=start_col + 1, end_row=7, end_column=start_col + 2
        )
        ws.cell(row=7, column=start_col + 1, value="Daily Target")
        n_days = (
            (last_data_row - first_data_row + 1)
            if last_data_row >= first_data_row
            else 0
        )
        ws.cell(
            row=7,
            column=start_col + daily_target_col_idx,
            value=(
                f"=IFERROR(ROUND({col_d}6 / MAX(1, DATE({li.end_date.year},{li.end_date.month},{li.end_date.day}) - TODAY() + 1), 0), 0)"
                if n_days
                else 0
            ),
        )

        for r in range(4, 8):
            for c in range(start_col, start_col + daily_target_col_idx + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font if c == start_col + 1 else value_font
                if c == start_col + 1:
                    cell.fill = label_fill
                cell.border = thin_border
                cell.alignment = center_align

        ws.cell(row=8, column=start_col, value="Daily Report")
        ws.cell(
            row=8,
            column=start_col + 1,
            value=f"=SUM({impressions_col_letter}{first_data_row}:{impressions_col_letter}{last_data_row})",
        )
        clicks_col_letter = col_letters[2]
        ws.cell(
            row=8,
            column=start_col + 2,
            value=f"=SUM({clicks_col_letter}{first_data_row}:{clicks_col_letter}{last_data_row})",
        )
        ctr_target_col = start_col + 3
        ws.cell(
            row=8,
            column=ctr_target_col,
            value=f"=IFERROR(({clicks_col_letter}8/{impressions_col_letter}8),0)",
        )
        for c in range(start_col, start_col + daily_target_col_idx + 1):
            cell = ws.cell(row=8, column=c)
            cell.font = header_font
            cell.fill = label_fill
            cell.border = thin_border
            cell.alignment = center_align

    # ── 1) Non-video line items: side-by-side narrow blocks on the main sheet (unchanged behavior) ──
    BLOCK_WIDTH = 6
    col_cursor = 1

    for li in non_video_line_items:
        start_col = col_cursor
        col_a = get_column_letter(start_col)
        col_b = get_column_letter(start_col + 1)
        col_c = get_column_letter(start_col + 2)
        col_d = get_column_letter(start_col + 3)

        ws_main.column_dimensions[col_a].width = 16
        ws_main.column_dimensions[col_b].width = 14
        ws_main.column_dimensions[col_c].width = 12
        ws_main.column_dimensions[col_d].width = 10
        for spacer_offset in range(4, BLOCK_WIDTH):
            ws_main.column_dimensions[
                get_column_letter(start_col + spacer_offset)
            ].width = 10

        if not li.start_date or not li.end_date:
            col_cursor += BLOCK_WIDTH
            continue

        days = list(daterange(li.start_date, li.end_date))
        n_days = len(days)
        target_impressions = li.impressions or 0
        first_data_row = 10
        last_data_row = first_data_row + n_days - 1

        write_summary_header(
            ws_main,
            start_col,
            li,
            target_impressions,
            first_data_row,
            last_data_row,
            col_letters=[col_a, col_b, col_c, col_d],
            daily_target_col_idx=3,
        )

        ws_main.cell(row=9, column=start_col, value="Date")
        ws_main.cell(row=9, column=start_col + 1, value="Impressions")
        ws_main.cell(row=9, column=start_col + 2, value="Clicks")
        ws_main.cell(row=9, column=start_col + 3, value="CTR")
        for c in range(start_col, start_col + 4):
            cell = ws_main.cell(row=9, column=c)
            cell.font = header_font
            cell.fill = label_fill
            cell.alignment = center_align
            cell.border = thin_border

        li_entries = entries_map.get(li.id, {})
        for idx, day in enumerate(days):
            row = first_data_row + idx
            entry = li_entries.get(day)
            imp_val = entry.impressions if entry else 0
            clk_val = entry.clicks if entry else 0
            ctr_val = entry.ctr if entry else 0

            ws_main.cell(row=row, column=start_col, value=day.strftime("%d %B, %Y"))
            ws_main.cell(row=row, column=start_col + 1, value=imp_val)
            ws_main.cell(row=row, column=start_col + 2, value=clk_val)
            ws_main.cell(row=row, column=start_col + 3, value=ctr_val)

            for c in range(start_col, start_col + 4):
                cell = ws_main.cell(row=row, column=c)
                cell.font = value_font
                cell.border = thin_border
                cell.alignment = center_align if c != start_col else left_align

        col_cursor += BLOCK_WIDTH

    # If there were no non-video line items at all, the main sheet stays empty — that's fine,
    # openpyxl allows an empty sheet to exist.

    # ── 2) Video line items: one dedicated wide sheet each ──
    VIDEO_COLS = [
        "Date",
        "Impressions",
        "Clicks",
        "CTR",
        "Start Views",
        "Complete Views",
        "Video Completion Rate",
        "Viewable Impressions",
        "Measurable Impressions",
        "Viewability",
    ]
    VIDEO_COL_WIDTHS = [16, 14, 12, 10, 12, 14, 22, 18, 20, 14]

    for li in video_line_items:
        # sheet_title = _safe_sheet_name(f"{li.line_item_name or li.line_item_id} - video", used_sheet_names)
        sheet_title = _safe_sheet_name(f"{li.line_item_id} - video", used_sheet_names)
        ws_v = wb.create_sheet(title=sheet_title)

        start_col = 1
        for offset, width in enumerate(VIDEO_COL_WIDTHS):
            ws_v.column_dimensions[get_column_letter(start_col + offset)].width = width

        if not li.start_date or not li.end_date:
            continue

        days = list(daterange(li.start_date, li.end_date))
        n_days = len(days)
        target_impressions = li.impressions or 0
        first_data_row = 10
        last_data_row = first_data_row + n_days - 1

        col_letters = [get_column_letter(start_col + i) for i in range(len(VIDEO_COLS))]
        # daily_target_col_idx mirrors the narrow-block layout (index 3 = the "CTR" / 4th column)
        # so the summary block (rows 3-8) still lines up visually under the first 4 columns.
        write_summary_header(
            ws_v,
            start_col,
            li,
            target_impressions,
            first_data_row,
            last_data_row,
            col_letters=col_letters,
            daily_target_col_idx=3,
        )

        for i, col_name in enumerate(VIDEO_COLS):
            cell = ws_v.cell(row=9, column=start_col + i, value=col_name)
            cell.font = header_font
            cell.fill = label_fill
            cell.alignment = center_align
            cell.border = thin_border

        li_entries = entries_map.get(li.id, {})
        for idx, day in enumerate(days):
            row = first_data_row + idx
            entry = li_entries.get(day)

            row_values = [
                day.strftime("%d %B, %Y"),
                entry.impressions if entry else 0,
                entry.clicks if entry else 0,
                entry.ctr if entry else 0,
                entry.video_start if entry else 0,  # Start Views
                entry.video_end if entry else 0,  # Complete Views
                getattr(entry, "vcr", 0) if entry else 0,  # VCR
                entry.viewable_impression if entry else 0,
                entry.measurable_impression if entry else 0,
                getattr(entry, "viewability", 0) if entry else 0,
            ]

            for i, val in enumerate(row_values):
                cell = ws_v.cell(row=row, column=start_col + i, value=val)
                cell.font = value_font
                cell.border = thin_border
                cell.alignment = center_align if i != 0 else left_align

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── NEW: Add a single daily entry for a line item (called from the frontend "Add" form) ──


@api_view(["POST"])
def add_daily_entry(request, campaign_id):
    """
    Body: { "line_item_id": "LIBILL002", "date": "2026-05-15", "impressions": 50000, "clicks": 1200, "ctr": 2.4 }
    Creates ONE row, never overwrites — if an entry already exists for that date,
    returns 409 so the frontend can disable that date's Add button.
    """
    try:
        campaign = (
            Campaign.objects.select_related("insertion_order")
            .prefetch_related("line_items")
            .get(campaign_id=campaign_id)
        )
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    line_item_id = request.data.get("line_item_id")
    date_str = request.data.get("date")
    impressions = request.data.get("impressions")
    clicks = request.data.get("clicks")
    ctr = request.data.get("ctr")

    # ✅ ADD THESE
    viewable_impression = request.data.get("viewable_impression")
    measurable_impression = request.data.get("measurable_impression")
    video_start = request.data.get("video_start")
    video_end = request.data.get("video_end")
    revenue = request.data.get("revenue")
    media_cost = request.data.get("media_cost")

    # ✅ NEW — only sent by frontend when ad format is video
    vcr = request.data.get("vcr")
    viewability = request.data.get("viewability")

    if not line_item_id or not date_str:
        return Response({"error": "line_item_id and date are required"}, status=400)

    try:
        line_item = LineItem.objects.get(line_item_id=line_item_id, campaign=campaign)
    except LineItem.DoesNotExist:
        return Response({"error": "LineItem not found in this campaign"}, status=404)

    entry_date = parse_date(date_str)
    if not entry_date:
        return Response({"error": "Invalid date"}, status=400)

    if entry_date < line_item.start_date or entry_date > line_item.end_date:
        return Response(
            {
                "error": f"Date must be between {line_item.start_date} and {line_item.end_date}"
            },
            status=400,
        )

    if CampaignDailyEntry.objects.filter(
        line_item=line_item, entry_date=entry_date
    ).exists():
        return Response(
            {"error": "Entry already exists for this date. Editing is not allowed."},
            status=409,
        )

    def safe_num(v, cast=int):
        try:
            return cast(v)
        except (TypeError, ValueError):
            return 0

    entry = CampaignDailyEntry.objects.create(
        campaign=campaign,
        line_item=line_item,
        entry_date=entry_date,
        impressions=safe_num(impressions, int),
        clicks=safe_num(clicks, int),
        ctr=safe_num(ctr, float),
        # ✅ ADD THESE
        viewable_impression=safe_num(viewable_impression, int),
        measurable_impression=safe_num(measurable_impression, int),
        video_start=safe_num(video_start, int),
        video_end=safe_num(video_end, int),
        revenue=safe_num(revenue, float),
        media_cost=safe_num(media_cost, float),
        # ✅ NEW
        vcr=safe_num(vcr, float),
        viewability=safe_num(viewability, float),
    )

    excel_bytes = build_daily_report_excel(campaign)
    filename = f"{campaign_id}_daily_report.xlsx"
    daily_excel_obj, _ = CampaignDailyReportExcel.objects.get_or_create(
        campaign=campaign
    )
    daily_excel_obj.excel_file.save(filename, ContentFile(excel_bytes), save=True)

    return Response(
        {
            "message": "Daily entry added successfully",
            "entry_id": entry.id,
            "line_item_id": line_item_id,
            "date": str(entry_date),
            "download_url": request.build_absolute_uri(daily_excel_obj.excel_file.url),
        },
        status=201,
    )


# ── NEW: Get already-submitted dates for a line item (to disable Add button for those dates) ──


@api_view(["GET"])
def get_daily_entries(request, campaign_id):
    line_item_id = request.query_params.get("line_item_id")

    try:
        campaign = Campaign.objects.get(campaign_id=campaign_id)
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    qs = CampaignDailyEntry.objects.filter(campaign=campaign)
    if line_item_id:
        qs = qs.filter(line_item__line_item_id=line_item_id)

    data = [
        {
            "line_item_id": e.line_item.line_item_id,
            "date": str(e.entry_date),
            "impressions": e.impressions,
            "clicks": e.clicks,
            "ctr": e.ctr,
            # ✅ ADD THESE
            "viewable_impression": e.viewable_impression,
            "measurable_impression": e.measurable_impression,
            "video_start": e.video_start,
            "video_end": e.video_end,
            "revenue": e.revenue,
            "media_cost": e.media_cost,
            # ✅ NEW
            "vcr": e.vcr,
            "viewability": e.viewability,
        }
        for e in qs.order_by("entry_date")
    ]
    return Response(data)


# ── NEW: Generate/regenerate the daily-report Excel on demand ──


@api_view(["POST"])
def generate_daily_report_excel(request, campaign_id):
    try:
        campaign = (
            Campaign.objects.select_related("client", "insertion_order")
            .prefetch_related("line_items")
            .get(campaign_id=campaign_id)
        )
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    excel_bytes = build_daily_report_excel(campaign)
    filename = f"{campaign_id}_daily_report.xlsx"

    daily_excel_obj, _ = CampaignDailyReportExcel.objects.get_or_create(
        campaign=campaign
    )
    daily_excel_obj.excel_file.save(filename, ContentFile(excel_bytes), save=True)

    return Response(
        {
            "message": "Daily report Excel generated successfully",
            "campaign_id": campaign_id,
            "download_url": request.build_absolute_uri(daily_excel_obj.excel_file.url),
            "filename": filename,
        },
        status=200,
    )


# ── NEW: Download the daily-report Excel ──


@api_view(["GET"])
def download_daily_report_excel(request, campaign_id):
    try:
        daily_excel_obj = CampaignDailyReportExcel.objects.get(
            campaign__campaign_id=campaign_id
        )
    except CampaignDailyReportExcel.DoesNotExist:
        try:
            campaign = (
                Campaign.objects.select_related("client", "insertion_order")
                .prefetch_related("line_items")
                .get(campaign_id=campaign_id)
            )
        except Campaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)

        excel_bytes = build_daily_report_excel(campaign)
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{campaign_id}_daily_report.xlsx"'
        )
        return response

    response = FileResponse(
        daily_excel_obj.excel_file.open("rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{campaign_id}_daily_report.xlsx"'
    )
    return response


import openpyxl
from rest_framework.parsers import MultiPartParser


@api_view(["POST"])
@parser_classes([MultiPartParser])
def bulk_upload_daily_entries(request):
    """
    Upload an Excel/CSV file with columns (header names, case-insensitive):
      campaign, line item id (this actually holds the dv_id value), date,
      impressions, clicks, click rate (ctr), revenue (adv currency),
      media cost (advertiser currency), start views, complete views,
      viewable impressions, measurable impressions

    Validates each row and inserts into CampaignDailyEntry.
    Regenerates daily report Excel for each affected campaign.
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        return Response({"error": "Invalid Excel file"}, status=400)

    # ── Map real-world header names → our internal keys ──
    # Keys on the right are what get_val() will be called with below.
    HEADER_MAP = {
        "campaign": "campaign_id",
        "line item id": "dv_id",  # ⚠️ this column actually holds dv_id values
        "date": "date",
        "impressions": "impressions",
        "clicks": "clicks",
        "click rate (ctr)": "ctr",
        "revenue (adv currency)": "revenue",
        "media cost (advertiser currency)": "media_cost",
        "start views": "video_start",
        "complete views": "video_end",
        "viewable impressions": "viewable_impression",
        "measurable impressions": "measurable_impression",
    }

    raw_headers = [
        str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]
    ]

    # Build col_idx keyed by OUR internal names, using HEADER_MAP to translate
    col_idx = {}
    for i, h in enumerate(raw_headers):
        internal_key = HEADER_MAP.get(h)
        if internal_key:
            col_idx[internal_key] = i

    required_keys = ["campaign_id", "dv_id", "date", "impressions", "clicks", "ctr"]
    missing = [k for k in required_keys if k not in col_idx]
    if missing:
        # Show the human-readable header name in the error, not our internal key
        reverse_map = {v: k for k, v in HEADER_MAP.items()}
        missing_labels = [reverse_map.get(k, k) for k in missing]
        return Response(
            {"error": f"Missing required column(s): {', '.join(missing_labels)}"},
            status=400,
        )

    def get_val(row, key):
        idx = col_idx.get(key)
        if idx is None:
            return None
        cell = row[idx]
        return cell.value

    results = {
        "inserted": [],
        "skipped": [],
        "errors": [],
    }

    affected_campaigns = {}  # campaign_id → campaign object

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Skip completely empty rows
        if all(cell.value is None for cell in row):
            continue

        campaign_id = str(get_val(row, "campaign_id") or "").strip()
        dv_id = str(get_val(row, "dv_id") or "").strip()
        date_raw = get_val(row, "date")
        impressions_raw = get_val(row, "impressions")
        clicks_raw = get_val(row, "clicks")
        ctr_raw = get_val(row, "ctr")
        revenue_raw = get_val(row, "revenue")
        media_cost_raw = get_val(row, "media_cost")
        video_start_raw = get_val(row, "video_start")
        video_end_raw = get_val(row, "video_end")
        viewable_imp_raw = get_val(row, "viewable_impression")
        measurable_imp_raw = get_val(row, "measurable_impression")

        # ── Validate required fields ──
        if not campaign_id or not dv_id or not date_raw:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": "Campaign, Line Item ID (DV ID), and Date are required",
                    "data": f"{campaign_id} / {dv_id} / {date_raw}",
                }
            )
            continue

        # ── Parse date ──
        try:
            if hasattr(date_raw, "date"):
                entry_date = date_raw.date()  # already a datetime from openpyxl
            else:
                from datetime import datetime as dt

                entry_date = dt.strptime(str(date_raw).strip(), "%Y-%m-%d").date()
        except Exception:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Invalid date format: '{date_raw}'. Use YYYY-MM-DD",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue

        # ── Validate campaign exists and is approved ──
        from campaigns.models import Campaign, LineItem

        try:
            campaign = Campaign.objects.get(
                campaign_id=campaign_id, approval_status="approved"
            )
        except Campaign.DoesNotExist:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Campaign '{campaign_id}' not found or not approved",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue

        # ── Validate line item belongs to campaign, looked up by dv_id ──
        try:
            line_item = LineItem.objects.get(dv_id=dv_id, campaign=campaign)
        except LineItem.DoesNotExist:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Line item with DV ID '{dv_id}' not found in campaign '{campaign_id}'",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue
        except LineItem.MultipleObjectsReturned:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Multiple line items share DV ID '{dv_id}' in campaign '{campaign_id}' — cannot determine which one",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue

        # ── Validate date is within line item range ──
        if line_item.start_date and entry_date < line_item.start_date:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Date {entry_date} is before line item start date {line_item.start_date}",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue

        if line_item.end_date and entry_date > line_item.end_date:
            results["errors"].append(
                {
                    "row": row_num,
                    "reason": f"Date {entry_date} is after line item end date {line_item.end_date}",
                    "data": f"{campaign_id} / {dv_id}",
                }
            )
            continue

        # ── Check duplicate ──
        if CampaignDailyEntry.objects.filter(
            line_item=line_item, entry_date=entry_date
        ).exists():
            results["skipped"].append(
                {
                    "row": row_num,
                    "reason": "Entry already exists for this date",
                    "data": f"{campaign_id} / {dv_id} / {entry_date}",
                }
            )
            continue

        # ── Safe number parsing ──
        def safe_int(v):
            try:
                return int(float(str(v).replace(",", "").strip()))
            except Exception:
                return 0

        def safe_float(v):
            try:
                return float(str(v).replace(",", "").replace("%", "").strip())
            except Exception:
                return 0.0

        # ── Insert entry, now with ALL fields ──
        CampaignDailyEntry.objects.create(
            campaign=campaign,
            line_item=line_item,
            entry_date=entry_date,
            impressions=safe_int(impressions_raw),
            clicks=safe_int(clicks_raw),
            ctr=safe_float(ctr_raw),
            viewable_impression=safe_int(viewable_imp_raw),
            measurable_impression=safe_int(measurable_imp_raw),
            video_start=safe_int(video_start_raw),
            video_end=safe_int(video_end_raw),
            revenue=safe_float(revenue_raw),
            media_cost=safe_float(media_cost_raw),
        )

        results["inserted"].append(
            {
                "row": row_num,
                "data": f"{campaign_id} / {dv_id} / {entry_date}",
            }
        )

        # ── Track affected campaigns for Excel regeneration ──
        affected_campaigns[campaign_id] = campaign

    # ── Regenerate daily report Excel for each affected campaign ──
    regenerated = []
    for cid, campaign in affected_campaigns.items():
        try:
            campaign_full = (
                Campaign.objects.select_related("insertion_order")
                .prefetch_related("line_items")
                .get(campaign_id=cid)
            )
            excel_bytes = build_daily_report_excel(campaign_full)
            filename = f"{cid}_daily_report.xlsx"
            daily_excel_obj, _ = CampaignDailyReportExcel.objects.get_or_create(
                campaign=campaign_full
            )
            daily_excel_obj.excel_file.save(
                filename, ContentFile(excel_bytes), save=True
            )
            regenerated.append(cid)
        except Exception as e:
            pass  # Don't fail the whole response if Excel generation fails

    return Response(
        {
            "message": "Bulk upload complete",
            "inserted": len(results["inserted"]),
            "skipped": len(results["skipped"]),
            "errors": len(results["errors"]),
            "details": results,
            "excel_regenerated_for": regenerated,
        },
        status=200,
    )


from datetime import date, timedelta
from django.db.models import Sum, Max
from campaigns.models import Campaign, LineItem
from .models import CampaignDailyEntry

def _compute_pacing_for_line_item(li, today=None):
    today = today or date.today()

    totals = CampaignDailyEntry.objects.filter(line_item=li).aggregate(
        total_delivered=Sum("impressions"),
        total_clicks=Sum("clicks"),
    )
    total_delivered = totals["total_delivered"] or 0
    total_clicks = totals["total_clicks"] or 0

    last_entry = (
        CampaignDailyEntry.objects.filter(line_item=li)
        .order_by("-entry_date")
        .first()
    )
    last_date = last_entry.entry_date if last_entry else None
    last_day_impression = last_entry.impressions if last_entry else None

    target_impressions = li.impressions or 0
    remaining = max(target_impressions - total_delivered, 0)

    if li.end_date:
        days_left = max((li.end_date - today).days + 1, 1)
    else:
        days_left = 1
    daily_target = round(remaining / days_left)

    flight_active = bool(li.start_date and li.end_date and li.start_date <= today <= li.end_date)

    status = None
    pct = None

    if not flight_active:
        status = None
    elif last_date is None:
        # truly zero entries ever submitted — nothing to compare, skip it
        status = None
    elif daily_target > 0:
        pct = round(((last_day_impression - daily_target) / daily_target) * 100, 0)
        status = "under" if last_day_impression < daily_target else "over"
    else:
        status = "over"
        pct = 0

    return {
        "total_delivered": total_delivered,
        "total_clicks": total_clicks,
        "daily_target": daily_target,
        "last_date": last_date,
        "last_day_impression": last_day_impression,
        "status": status,
        "pct": pct,
    }
@api_view(["GET"])
def get_pacing_report(request):
    """
    GET /get_pacing_report/?status=under   (or over / not_uploaded)
    Omit ?status to get everything (used internally / for debugging).
    """
    status_filter = request.query_params.get("status")
    today = date.today()

    campaigns = (
        Campaign.objects.filter(approval_status="approved")
        .select_related("client", "insertion_order")
        .prefetch_related("line_items")
    )

    rows = []
    s_no = 1

    for campaign in campaigns:
        insertion_order = getattr(campaign, "insertion_order", None)

        for li in campaign.line_items.all():
            if not li.start_date or not li.end_date:
                continue

            pacing = _compute_pacing_for_line_item(li, today)

            if pacing["status"] is None:
                continue  # not active / no flight dates
            if status_filter and pacing["status"] != status_filter:
                continue

            rows.append({
                "s_no": s_no,
                "campaign_id": campaign.campaign_id,
                "campaign_name": campaign.campaign_name,
                "client_id": campaign.client.client_id if campaign.client else "",
                "client_name": campaign.client.name if campaign.client else "",
                "io_id": insertion_order.io_id if insertion_order else "",
                "io_name": campaign.campaign_name or "",
                "line_item_id": li.line_item_id,
                "line_item_name": li.line_item_name,
                "flight_dates": f"{li.start_date.strftime('%B %d, %Y')} - {li.end_date.strftime('%B %d, %Y')}",
                "total_volume_booked": li.impressions or 0,
                "total_volume_delivered": pacing["total_delivered"],
                "total_clicks": pacing["total_clicks"],
                "daily_target": pacing["daily_target"],
                "last_day_impression": pacing["last_day_impression"],
                "last_day_date": str(pacing["last_date"]) if pacing["last_date"] else None,
                "pct": pacing["pct"],
                "status": pacing["status"],
            })
            s_no += 1

    return Response(rows)