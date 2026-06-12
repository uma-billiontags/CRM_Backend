# ----- Import libraries ------

from django.http import HttpResponse
from rest_framework.decorators import api_view, parser_classes
from rest_framework.response import Response 
from rest_framework import status
from .models import Client, Campaign,LineItem,Creative, ThirdPartyCreative, Client, User, TeamAccess
from .serializers import ClientSerializer, CampaignSerializer, TeamAccessSerializer
import json
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Prefetch
from django.db import transaction  # imports Django transaction management system.
from datetime import datetime
from django.contrib.auth.hashers import make_password
from rest_framework.decorators import api_view
from django.contrib.auth import authenticate
from django.utils import timezone
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from .notification import send_push_notification



# ==============================
# Home function 
# ==============================
def home(request):
    return HttpResponse("Welcome to the CRM Home Page!")

# ==============================
# DEFINE CLIENT FUNCTION
# ==============================

from rest_framework.decorators import api_view
from rest_framework.response import Response

from .serializers import ClientSerializer
from .models import Client, User

from .notification import send_push_notification

import json


@api_view(['POST'])
def create_client(request):

    try:

        # ==========================
        # GET SIGNATURE FILES
        # ==========================

        signatures = {

            key: request.FILES[key]

            for key in request.FILES

            if key.startswith(
                'contact_signature_'
            )
        }

        # ==========================
        # GET FORM DATA
        # ==========================

        raw = request.data.get('data')

        if raw:

            parsed = json.loads(raw)

            data = parsed

        else:

            data = request.data

        # ==========================
        # CHECK DUPLICATE EMAIL
        # ==========================

        email = data.get("email")

        if Client.objects.filter(
            email=email
        ).exists():

            return Response({

                "error":
                "This email is already registered"

            }, status=400)

        # ==========================
        # SERIALIZER VALIDATION
        # ==========================

        serializer = ClientSerializer(

            data=data,

            context={
                'signatures': signatures
            }
        )

        if serializer.is_valid():

            # ==========================
            # SAVE CLIENT
            # ==========================

            client = serializer.save()

            # ==========================
            # CREATE LOGIN USER
            # ==========================

            if email and not User.objects.filter(
                email=email
            ).exists():

                User.objects.create(

                    username=email,

                    email=email,

                    role='client',

                    client=client
                )

            # ==========================
            # SEND FIREBASE PUSH NOTIFICATION
            # ==========================

            send_push_notification(

                "New Client Request",

                f"New Client Submitted: {client.name}"

            )

            # ==========================
            # SUCCESS RESPONSE
            # ==========================

            return Response({

                "message":
                "Client created successfully"

            }, status=201)

        # ==========================
        # SERIALIZER ERROR RESPONSE
        # ==========================

        return Response(

            serializer.errors,

            status=400
        )

    except Exception as e:

        # ==========================
        # EXCEPTION ERROR RESPONSE
        # ==========================

        return Response({

            "error":
            str(e)

        }, status=500)





# To update the client function for approvel
@api_view(['PATCH'])
def update_client_status(request, client_id):
    try:
        client = Client.objects.get(client_id=client_id)
    except Client.DoesNotExist:
        return Response({"error": "Client not found"}, status=404)

    status = request.data.get('status')
    if status not in ['pending', 'approved', 'rejected']:
        return Response({"error": "Invalid status"}, status=400)

    client.status = status
    client.save()
    return Response({"message": f"Client status updated to {status}"}, status=200)






# ==============================
# GET SINGLE CLIENT
# ==============================
@api_view(['GET'])
def get_client(request, client_id):
    try:
        client = Client.objects.select_related(
            'billing',          #  FIXED
            'ownership',        #  FIXED
            'classification'    #  FIXED
        ).prefetch_related(
            'addresses',        # correct
            'contacts'          # correct
        ).get(client_id=client_id)

    except Client.DoesNotExist:
        return Response({"error": "Client not found"}, status=404)

    serializer = ClientSerializer(client)
    return Response(serializer.data)


# ==============================
# GET ALL CLIENTS
# ==============================

@api_view(['GET'])
def get_all_clients(request):

    clients = Client.objects.select_related(
        'billing',          # FIXED
        'ownership',        #  FIXED
        'classification'    #  FIXED
    ).prefetch_related(
        'addresses',        #  correct
        'contacts'          #  correct
    ).all()

    serializer = ClientSerializer(clients, many=True)
    return Response(serializer.data)


# ------------------------------------------------------------------------------------
# ==============================
# CREATE CAMPAIGN
# ==============================

# Creates a custom function to convert frontend date string into Python date.
def parse_date(date_str):

    if not date_str:
        return None
    return datetime.fromisoformat(date_str.replace('Z', '')).date()


@api_view(['POST']) 
@parser_classes([MultiPartParser, FormParser])  # it allows file upload, formdata. (without this file uploads won't work)
def create_campaign(request):

    client_id = request.data.get('client') # get the client id CLT-2026-00001 from frontend
    if not client_id:
        return Response({"error": "client is required"},status=400) # if frontend not send throw this error

    # =====================================================
    # FIND CLIENT
    # =====================================================

    try:
        client = Client.objects.get(client_id=client_id) # find the client(full details) from database
    except Client.DoesNotExist:
        return Response({"error": f"Client '{client_id}' not found"},status=404)

    # =====================================================
    # VALIDATE CAMPAIGN
    # =====================================================

    serializer = CampaignSerializer(data=request.data) # Take all data coming from frontend request and give it to serializer.
    if not serializer.is_valid():
        return Response(serializer.errors,status=400)

    try:
        with transaction.atomic(): # save together
            campaign = serializer.save(client=client) # save campagin

            try:
                line_items_data = json.loads(request.data.get('line_items','[]')) # get the line items from frontend and converts json into python dict
            except Exception:
                return Response({"error": "Invalid line_items JSON"},status=400)

            # =============================================
            # LOOP LINE ITEMS
            # =============================================

            for i, li in enumerate(line_items_data): # Loop through every line item one by one. (eg: i=0, li=first line item)

                line_item_id = li.get('line_item_id') # get the line item id (LIUSER0001)

                if not line_item_id:   # If no ID, skip that line item.
                    continue

                line_item, _ = LineItem.objects.update_or_create(  # Search LineItem using line_item_id (if line item data already exist (update) or (create))
                    line_item_id=line_item_id,
                    defaults={

                        'campaign': campaign,
                        'line_item_name': li.get('lineItemName'),
                        'ethnicity': li.get('ethnicity',[]),
                        'start_date': parse_date(li.get('startDate')),
                        'end_date': parse_date(li.get('endDate')),
                        'ad_format': li.get('adFormat',[]),
                        'impressions': li.get('impressions') or None,
                        'units': li.get('units') or None,
                        'ctr': li.get('ctr') or None,
                        'viewability': li.get('viewability') or None,
                        'vcr': li.get('vcr') or None,
                        'unit_cost': li.get('unit_cost') or None,
                        'kpi_notes': li.get('kpi_notes', ''),
                        'unit_value': li.get('unit_value') or None,
                        
                        'age':            li.get('age', ''),
                        'gender':         li.get('gender', ''),
                        'geo_targeting':  li.get('geo_targeting', ''),
                        'platforms':      li.get('platforms', ''),
                        'frequency_cap':  li.get('frequency_cap') or None,
                        'brand_safety':   li.get('brand_safety', ''),

                    }
                )

        
                # Fix — two separate loops

                # Loop 1: normal creatives
                creatives_meta = li.get('creatives', [])
                for j, meta in enumerate(creatives_meta):
                    if not meta.get('creative_name'):
                        continue

                    main_asset = request.FILES.get(f'line_item_{i}main_asset{j}')
                    Creative.objects.create(
                        line_item=line_item,
                        creative_name=meta.get('creative_name', ''),
                        main_asset=main_asset,
                        dimensions=meta.get('dimensions', ''),
                        aspect_ratio=meta.get('aspect_ratio', ''),
                        file_size=meta.get('file_size', ''),
                        click_through_url=meta.get('click_through_url') or None,
                        appended_html_tag=meta.get('appended_html_tag', ''),
                        integration_code=meta.get('integration_code', ''),
                        notes=meta.get('notes', ''),
                    )

                # Loop 2: third-party creatives — completely separate
                third_party_meta = li.get('third_party_creatives', [])
                for k, _ in enumerate(third_party_meta):
                    input_file = request.FILES.get(f'line_item_{i}thirdparty_file{k}')
                    backup_image = request.FILES.get(f'line_item_{i}thirdparty_backup{k}')
                    ThirdPartyCreative.objects.create(
                        line_item=line_item,
                        input_file=input_file,
                        backup_image=backup_image,
                    )
                
    except Exception as e:
        return Response({"error": str(e)},status=500)


    # =============================================
    # SEND FIREBASE PUSH NOTIFICATION
    # =============================================

    send_push_notification(
        "New Campaign Submitted",
        # f"Campaign {campaign.campaign_id} submitted for client {client.name}"
        f"{client.name} submitted the campaign for client ID {client.client_id}"
    )

    return Response({"message": (
            "Campaign + LineItem + "
            "Creative + ThirdPartyCreative "
            "saved successfully"
        ),
        "campaign_id": campaign.campaign_id,}, status=201)



#===========================
# GET ALL CAMPAIGNS
# ==========================

@api_view(['GET'])
def get_campaigns(request):

    campaigns = Campaign.objects.select_related('client').prefetch_related(
        'line_items__creatives_detail',
        'line_items__third_party_creatives' 
    )
    serializer = CampaignSerializer(campaigns, many=True)
    return Response(serializer.data)


# ==============================
# GET CAMPAIGNS BY CLIENT ID
# ==============================
    
@api_view(['GET'])
def get_campaigns_by_client(request, client_id):  # http://127.0.0.1:8000/get_campaigns_by_client/CLT-2026-00001/

    try:
        campaigns = Campaign.objects.filter(
            client__client_id=client_id
        ).select_related('client').prefetch_related(
            Prefetch(
                'line_items',
                queryset=LineItem.objects.prefetch_related('creatives_detail')
            )
        )

        if not campaigns.exists():
            return Response({"message": "No campaigns found for this client"}, status=404)

        serializer = CampaignSerializer(campaigns, many=True) 
        return Response(serializer.data)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

# ==============================
# GET CAMPAIGNS BY CAMPAIGN ID
# ==============================

@api_view(['GET'])
def get_campaign_by_id(request, campaign_id):  # http://127.0.0.1:8000/get_campaign_by_id/CMP-2026-00001/

    try:
        campaign = Campaign.objects.select_related('client').prefetch_related(
            Prefetch(
                'line_items',
                queryset=LineItem.objects.prefetch_related('creatives_detail')
            )
        ).get(campaign_id=campaign_id)

    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

    serializer = CampaignSerializer(campaign)
    return Response(serializer.data)


# -------------------------------------------------

# =========================================================
# UPDATE CAMPAIGN
# =========================================================

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
@api_view(['GET', 'PUT'])
@parser_classes([MultiPartParser, FormParser, JSONParser])

def update_campaign(request, campaign_id):

    # =====================================================
    # FIND CAMPAIGN
    # =====================================================

    try:

        campaign = Campaign.objects.select_related(
            'client'
        ).prefetch_related(
            'line_items__creatives_detail',
            'line_items__third_party_creatives'
        ).get(
            campaign_id=campaign_id
        )

    except Campaign.DoesNotExist:

        return Response(
            {"error": "Campaign not found"},
            status=404
        )

    # =====================================================
    # GET CAMPAIGN
    # =====================================================

    if request.method == 'GET':

        serializer = CampaignSerializer(
            campaign,
            context={'request': request}
        )

        return Response(serializer.data)

    # =====================================================
    # UPDATE CAMPAIGN
    # =====================================================

    try:

        with transaction.atomic():

            data = {}
            for key, value in request.data.items():
                if key != 'line_items':   # ← exclude line_items from campaign serializer
                    data[key] = value

            serializer = CampaignSerializer(

                campaign,

                data=request.data,

                partial=True,

                context={'request': request}
            )



            if not serializer.is_valid():

                return Response(
                    serializer.errors,
                    status=400
                )

            # =============================================
            # SAVE CAMPAIGN
            # =============================================

            serializer.save(

                new_cpm=(
                    request.data.get('new_cpm')
                    if request.data.get('new_cpm')
                    else None
                ),

                new_price=(
                    request.data.get('new_price')
                    if request.data.get('new_price')
                    else None
                )
            )

            # =============================================
            # GET LINE ITEMS
            # =============================================

            try:

                line_items_data = json.loads(
                    request.data.get(
                        'line_items',
                        '[]'
                    )
                )
                

            except Exception:

                return Response(
                    {"error": "Invalid line_items JSON"},
                    status=400
                )

            # =============================================
            # LOOP LINE ITEMS
            # =============================================

            for i, li in enumerate(line_items_data):

                line_item_id = li.get(
                    'line_item_id'
                )

                if not line_item_id:
                    continue

                line_item, _ = LineItem.objects.update_or_create(

                    line_item_id=line_item_id,

                    defaults={

                        'campaign': campaign,

                        'line_item_name':
                        li.get('lineItemName'),

                        'ethnicity':
                        li.get('ethnicity', []),

                        'start_date':
                        parse_date(li.get('startDate')),

                        'end_date':
                        parse_date(li.get('endDate')),

                        'ad_format':
                        li.get('adFormat', []),

                        'impressions':
                        li.get('impressions') or None,

                        'units':
                        li.get('units') or None,

                        'ctr':
                        li.get('ctr') or None,

                        'viewability':
                        li.get('viewability') or None,

                        'vcr':
                        li.get('vcr') or None,

                        'unit_cost':
                        li.get('unit_cost') or None,

                        'kpi_notes':
                        li.get('kpi_notes', ''),

                        'unit_value':
                        li.get('unit_value') or None,
                        
                        'age':            li.get('age', ''),
                        'gender':         li.get('gender', ''),
                        'geo_targeting':  li.get('geo_targeting', ''),
                        'platforms':      li.get('platforms', ''),
                        'frequency_cap':  li.get('frequency_cap') or None,
                        'brand_safety':   li.get('brand_safety', ''),
            }
                )

                # =============================================
                # NORMAL CREATIVES
                # =============================================

                creatives_meta = li.get(
                    'creatives',
                    []
                )

                for j, meta in enumerate(creatives_meta):

                    if not meta.get(
                        'creative_name'
                    ):
                        continue

                    main_asset = request.FILES.get(
                        f'line_item_{i}main_asset{j}'
                    )

                    Creative.objects.create(

                        line_item=line_item,

                        creative_name=meta.get(
                            'creative_name',
                            ''
                        ),

                        main_asset=main_asset,

                        dimensions=meta.get(
                            'dimensions',
                            ''
                        ),

                        aspect_ratio=meta.get(
                            'aspect_ratio',
                            ''
                        ),

                        file_size=meta.get(
                            'file_size',
                            ''
                        ),

                        click_through_url=meta.get(
                            'click_through_url'
                        ) or None,

                        appended_html_tag=meta.get(
                            'appended_html_tag',
                            ''
                        ),

                        integration_code=meta.get(
                            'integration_code',
                            ''
                        ),

                        notes=meta.get(
                            'notes',
                            ''
                        ),
                    )

                # =============================================
                # THIRD PARTY CREATIVES
                # =============================================

                third_party_meta = li.get(
                    'third_party_creatives',
                    []
                )

                for k, _ in enumerate(third_party_meta):

                    input_file = request.FILES.get(
                        f'line_item_{i}thirdparty_file{k}'
                    )

                    backup_image = request.FILES.get(
                        f'line_item_{i}thirdparty_backup{k}'
                    )

                    ThirdPartyCreative.objects.create(

                        line_item=line_item,

                        input_file=input_file,

                        backup_image=backup_image,
                    )

    except Exception as e:

        import traceback

        print(traceback.format_exc())

        return Response(
            {"error": str(e)},
            status=500
        )

    # =====================================================
    # SUCCESS RESPONSE
    # =====================================================

    return Response({

        "message":
        "Campaign updated successfully",

        "campaign_id":
        campaign.campaign_id,

    }, status=200)


# =========================================================
# DELETE CAMPAIGN
# =========================================================

@api_view(['DELETE'])
def delete_campaign(request, campaign_id):

    # =====================================================
    # FIND CAMPAIGN
    # =====================================================

    try:

        # ✅ Try campaign_id string first, fallback to pk
        try:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            # For pending campaigns where campaign_id is null, try pk
            campaign = Campaign.objects.get(pk=campaign_id)

    except Campaign.DoesNotExist:
        

        return Response({

            "error": "Campaign not found"

        }, status=404)

    # =====================================================
    # DELETE CAMPAIGN
    # =====================================================

    campaign.delete()

    # =====================================================
    # SUCCESS RESPONSE
    # =====================================================

    return Response({

        "message": "Campaign deleted successfully",

        "campaign_id": campaign_id

    }, status=200)




# ------------- Login functionality ---------------

@api_view(['POST'])
def login_view(request):
    email = request.data.get('email') # get the email and password from frontend request body
    password = request.data.get('password')

    if not email or not password:
        return Response({"error": "Email and password are required"}, status=400) 

    # ── 1. Check TeamAccess table first (team members) ──────────────────────
    try:
        team_member = TeamAccess.objects.get(email=email)

        if team_member.status != 'Active':
            return Response({"error": "Your account is inactive. Contact your administrator."}, status=403)

        if team_member.password != password:
            return Response({"error": "Invalid password"}, status=401)
        
        role = team_member.role.lower().replace(" ", "_")
        
        return Response({
            "message": "Login successful",
            "user": {
                "id":        team_member.id,
                "username":  team_member.member,
                "email":     team_member.email,
                "role":      role,
                "client_id": None,
                "source":    "team",
            }
        }, status=200)

    except TeamAccess.DoesNotExist:
        pass  # not a team member, check user table next

    # ── 2. Check User table (clients + superadmin) ───────────────────────────
    try:
        user_obj = User.objects.get(email=email)   # searches the user in db using email
    except User.DoesNotExist:
        return Response({"error": "No account found with this email"}, status=401)

    user = authenticate(username=user_obj.username, password=password)   # verifies the password & username using django built-in authentication system
    if user is None:
        return Response({"error": "Invalid password"}, status=401)
    
    # UPDATE LAST LOGIN TIME    
    user.last_login = timezone.now()
    user.save()
    
    if user.role == 'client':   # check if the logged user is client
        if user.client is None or user.client.status != 'approved':      # check the client approvel status
            return Response({"error": "Your account is not approved yet"}, status=403)

    return Response({
        "message": "Login successful",
        "user": {
            "id":        user.id,
            "username":  user.username,
            "email":     user.email,
            "role":      user.role,
            "client_id": user.client.client_id if user.client else None,     # If user linked with client: return client id otherwise return None
            "client_name": user.client.name if user.client else None,  # ← ADD THIS
            "source":    "user",
        }
    }, status=200)


# ------------------ Download function ----------------------

from django.http import FileResponse
from django.shortcuts import get_object_or_404
import mimetypes

@api_view(['GET'])
def download_creative(request, creative_id):

    # Get creative object
    creative = get_object_or_404(Creative,id=creative_id)

    # Get uploaded file path
    file_path = creative.main_asset.path
    
    mime_type, _ = mimetypes.guess_type(file_path)
    mime_type = mime_type or 'application/octet-stream'


    # Return downloadable response
    return FileResponse(open(file_path, 'rb'),as_attachment=True,content_type=mime_type) 


# Third party function
@api_view(['GET'])
def download_thirdparty(request,thirdparty_id):

    # Get third-party creative object
    thirdparty = get_object_or_404(ThirdPartyCreative,id=thirdparty_id)

    if not thirdparty.input_file:
        return Response({"error":"No input file uploaded"},status=404)

    # Get uploaded file path
    file_path = thirdparty.input_file.path

    # Return downloadable response
    return FileResponse(open(file_path, 'rb'),as_attachment=True,filename=thirdparty.input_file.name)


# Backup image function
@api_view(['GET'])
def download_backup_image(request,thirdparty_id):

    thirdparty = get_object_or_404(ThirdPartyCreative,id=thirdparty_id)

    if not thirdparty.backup_image:
        return Response({"error":"No input file uploaded"},status=404)
    file_path = thirdparty.backup_image.path
    return FileResponse(open(file_path, 'rb'),as_attachment=True,filename=thirdparty.backup_image.name)


# ===============================
# APPROVAL FUNCTIONALITY
# ===============================
from django.core.mail import send_mail

@api_view(['POST'])
def approve_client(request):

    # =====================================
    # GET DATA FROM FRONTEND
    # =====================================

    client_id = request.data.get('client_id')

    password = request.data.get('password')

    # =====================================
    # VALIDATION
    # =====================================

    if not client_id or not password:

        return Response({

            "error":
            "client_id and password required"

        }, status=400)

    # =====================================
    # GET CLIENT
    # =====================================
    try:

        client = Client.objects.get(
            client_id=client_id
        )

    except Client.DoesNotExist:

        return Response({

            "error":
            "Client not found"

        }, status=404)

    # =====================================
    # CREATE USER IF NOT EXISTS
    # =====================================

    user, created = User.objects.get_or_create(

        email=client.email,

        defaults={

            "username": client.client_id,

            "email": client.email,

            "role": "client",

            "client": client
        }
    )

    # =====================================
    # UPDATE USER DETAILS
    # =====================================

    user.username = client.email   # client.email  (or) client.client_id

    user.email = client.email

    user.role = "client"

    user.client = client

    # SET PASSWORD (HASH PASSWORD)
    user.set_password(password)

    # SAVE USER
    user.save()

    # =====================================
    # UPDATE CLIENT STATUS
    # =====================================

    client.status = "approved"

    client.save()

    # =====================================
    # SEND LOGIN EMAIL
    # =====================================

    send_mail(

        subject='CRM Login Credentials',

        message=f'''
Hello {client.name},

Your account has been approved.

Login Email:
{client.email}

Password:
{password}

You can now login and create campaigns.
Thank You
''',

        from_email='yourgmail@gmail.com',

        recipient_list=[client.email],

        fail_silently=False
    )

    # =====================================
    # SUCCESS RESPONSE
    # =====================================

    return Response({

        "message":
        "Client approved successfully",

        "client_id":
        client.client_id,

        "email":
        client.email,

        "status":
        client.status

    }, status=200)


# ---------------------------
# TEAM MEMBER MANAGEMENT
# ---------------------------

# Create Team Member
@api_view(['POST'])
def create_team_member(request):

    serializer = TeamAccessSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Get All Team Members
@api_view(['GET'])
def get_team_members(request):

    members = TeamAccess.objects.all().order_by('-id')
    serializer = TeamAccessSerializer(members, many=True)

    return Response(serializer.data)


# Edit Team Member
@api_view(['PUT'])
def edit_team_member(request, id):

    try:
        member = TeamAccess.objects.get(id=id)
    except TeamAccess.DoesNotExist:
        return Response(
            {"error": "Team member not found"},
            status=status.HTTP_404_NOT_FOUND
        )

    serializer = TeamAccessSerializer(member, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Delete Team Member
@api_view(['DELETE'])
def delete_team_member(request, id):

    try:
        member = TeamAccess.objects.get(id=id)
    except TeamAccess.DoesNotExist:
        return Response(
            {"error": "Team member not found"},
            status=status.HTTP_404_NOT_FOUND
        )

    member.delete()

    return Response(
        {"message": "Team member deleted successfully"},
        status=status.HTTP_200_OK
    )


# ---------------------------
# USER MANAGEMENT
# ---------------------------

# GET all client users
@api_view(['GET'])
def get_client_users(request):
    users = User.objects.filter(role='client').select_related('client')
    data = []
    for u in users:
        data.append({
            "id": u.id,
            "client_id": u.client.client_id if u.client else u.username,
            "email": u.email,
            "role": u.role,
            "status": getattr(u, 'status', 'Active'),  # once you add status field
            "last_active": u.last_login.isoformat() if u.last_login else u.date_joined.isoformat(),
        })
    return Response(data)



@api_view(['DELETE'])
def delete_client_user(request, id):

    # =====================================
    # FIND CLIENT USER
    # =====================================

    try:

        user = User.objects.select_related(
            'client'
        ).get(

            id=id,

            role='client'
        )

    except User.DoesNotExist:

        return Response({

            "error":
            "Client user not found"

        }, status=404)

    # =====================================
    # STORE DETAILS
    # =====================================

    client = user.client

    deleted_data = {

        "user_id": user.id,

        "email": user.email,

        "client_id": (

            client.client_id
            if client else None
        )
    }

    # =====================================
    # DELETE CLIENT TABLE
    # =====================================

    if client:

        client.delete()

    # =====================================
    # SUCCESS RESPONSE
    # =====================================

    return Response({

        "message":
        "Client and user deleted successfully",

        "deleted_data":
        deleted_data

    }, status=200)


@api_view(['PUT'])
def edit_client_user(request, id):
    try:
        user = User.objects.get(id=id)
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=404)

    password = request.data.get('password')
    status = request.data.get('status')

    if password:
        user.set_password(password)
    if status:
        user.status = status   # add this field to User model first

    user.save()
    return Response({"message": "Updated successfully"}, status=200)


# ---------------------------
# Approve campaign function
# ---------------------------
from .notification import send_notification_to_client

@api_view(['POST'])
def approve_campaign(request, pk):
    try:
        campaign = Campaign.objects.get(pk=pk)
    except Campaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)
    
    # ✅ Check approval_status instead of campaign_id
    if campaign.approval_status == 'approved':
        return Response({"error": "Campaign already approved"}, status=400)

    # if campaign.campaign_id:
    #     return Response({"error": "Campaign already approved"}, status=400)

    try:
        with transaction.atomic():
            new_id = campaign.generate_campaign_id()

            # Extra safety: ensure uniqueness
            counter = 1
            original_id = new_id
            while Campaign.objects.filter(campaign_id=new_id).exists():
                # Increment last number manually
                parts = new_id.split('-')
                num = int(parts[-1]) + counter
                new_id = f"{'-'.join(parts[:-1])}-{str(num).zfill(5)}"
                counter += 1

            # Assign and save
            campaign.campaign_id = new_id
            campaign.approval_status = 'approved'
            campaign.save()   # This is critical

            # Send notification
            send_notification_to_client(
                campaign.client,
                "Campaign Approved",
                f"Your campaign {campaign.campaign_name} has been approved with ID: {new_id}"
            )

            print(f"✅ Campaign approved successfully: {new_id}")  # For debugging

            return Response({
                "message": "Campaign approved successfully",
                "campaign_id": campaign.campaign_id,
            }, status=200)

    except Exception as e:
        import traceback
        print("❌ Approve Campaign Error:")
        print(traceback.format_exc())
        return Response({
            "error": str(e),
            "details": "Check server logs for full traceback"
        }, status=500)

# from .notification import send_push_notification_to_user
# from .models import User

# @api_view(['POST'])
# def approve_campaign(request, pk):

#     try:
#         campaign = Campaign.objects.get(pk=pk)

#     except Campaign.DoesNotExist:
#         return Response(
#             {"error": "Campaign not found"},
#             status=404
#         )

#     if campaign.campaign_id:
#         return Response(
#             {"error": "Campaign already approved"},
#             status=400
#         )

#     for i in range(5):

#         with transaction.atomic():

#             new_id = campaign.generate_campaign_id()

#             if not Campaign.objects.filter(
#                 campaign_id=new_id
#             ).exists():

#                 campaign.campaign_id = new_id
#                 campaign.approval_status = 'approved'
#                 campaign.save()

#                 # ==================================
#                 # SEND NOTIFICATION TO CLIENT
#                 # ==================================

#                 try:

#                     client_user = User.objects.get(
#                         client=campaign.client
#                     )

#                     send_push_notification_to_user(

#                         client_user,

#                         "Campaign Approved",

#                         f"Your campaign {campaign.campaign_id} has been approved."

#                     )

#                 except User.DoesNotExist:

#                     print(
#                         f"No user found for client "
#                         f"{campaign.client.name}"
#                     )

#                 break

#     return Response(
#         {
#             "message": "Campaign approved",
#             "campaign_id": campaign.campaign_id,
#         },
#         status=200
#     )



# ==========================================
# FCM Token Update Function
# ==========================================

from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
import json
from .models import FCMToken, User

# working perfectly

# @api_view(['POST'])
# @csrf_exempt
# def save_fcm_token(request):
#     print('User:', request.user)  # Check if user is authenticated

#     try:

#         data = json.loads(request.body)

#         token = data.get("token")
    

#         if token:

#             FCMToken.objects.get_or_create(token=token)

#             return Response({"message":"Token Saved Successfully"}, status=200)

#         return Response({"error":"Token Missing"}, status=400)

#     except Exception as e:

#         return Response({"error":str(e)}, status=500)




# @api_view(['POST'])
# @csrf_exempt
# def save_fcm_token(request):

#     try:

#         data = json.loads(request.body)

#         token = data.get("token")
#         email = data.get("email")

#         if not token:
#             return Response(
#                 {"error": "Token Missing"},
#                 status=400
#             )

#         user = None

#         if email:

#             try:
#                 user = User.objects.get(
#                     email=email
#                 )

#             except User.DoesNotExist:
#                 pass

#         FCMToken.objects.update_or_create(

#             token=token,

#             defaults={
#                 "user": user
#             }
#         )

#         return Response(
#             {
#                 "message":
#                 "Token Saved Successfully"
#             },
#             status=200
#         )

#     except Exception as e:

#         return Response(
#             {
#                 "error": str(e)
#             },
#             status=500
#         )



@api_view(['POST'])
@csrf_exempt
def save_fcm_token(request):
    try:
        data = json.loads(request.body)
        token = data.get("token")
        client_id = data.get("client_id")  # ← change from email to client_id

        if not token:
            return Response({"error": "Token Missing"}, status=400)

        #user = None
        client = None
        if client_id:
            try:
                # Find the User linked to this client_id
                from .models import Client  # adjust import to your app
                client = Client.objects.get(client_id=client_id)
                #user = User.objects.get(client=client)
            except (Client.DoesNotExist):
                pass

        FCMToken.objects.update_or_create(
            token=token,
            defaults={"client": client}
        )

        return Response({"message": "Token Saved Successfully"}, status=200)

    except Exception as e:
        return Response({"error": str(e)}, status=500)




# ==========================================


# Claude

# @api_view(['POST'])
# @csrf_exempt
# def save_fcm_token(request):

#     try:

#         data = json.loads(request.body)
#         token = data.get("token")

#         if not token:
#             return Response({"error": "Token Missing"}, status=400)

#         # Duplicate check — already exists skip
#         obj, created = FCMToken.objects.get_or_create(token=token)

#         if created:
#             print(f"New token saved: {token[:20]}...")
#         else:
#             print(f"Token already exists, skip")

#         return Response({"message": "Token Saved Successfully"}, status=200)

#     except Exception as e:
#         return Response({"error": str(e)}, status=500)

# ===============================================

# Claude

# ==============================
# CHAT HISTORY API
# ==============================

from .models import ChatRoom, Message
from .serializers import MessageSerializer

@api_view(['GET'])
def get_chat_history(request, campaign_id):

    try:

        # ==========================
        # FIND CHAT ROOM
        # ==========================

        try:
            room = ChatRoom.objects.get(
                campaign__campaign_id=campaign_id
            )
        except ChatRoom.DoesNotExist:
            return Response([], status=200)

        # ==========================
        # GET MESSAGES
        # ==========================

        messages = Message.objects.filter(
            room=room
        ).select_related(
            'sender'
        ).order_by('timestamp')  # oldest first → WhatsApp style

        serializer = MessageSerializer(
            messages,
            many=True,
            context={'request': request}  # ← ADD THIS LINE
        )

        return Response(serializer.data, status=200)

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )


# ==============================
# MARK MESSAGES AS READ
# ==============================

@api_view(['POST'])
def mark_messages_read(request, campaign_id):

    try:

        room = ChatRoom.objects.get(
            campaign__campaign_id=campaign_id
        )

        Message.objects.filter(
            room=room,
            is_read=False,
            sender_type='client'
        ).update(is_read=True)

        return Response(
            {"message": "Marked as read"},
            status=200
        )

    except ChatRoom.DoesNotExist:
        return Response(
            {"error": "Room not found"},
            status=404
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )


# ==============================
# GET ALL CHAT ROOMS (Admin view)
# ==============================

@api_view(['GET'])
def get_all_chat_rooms(request):

    try:

        rooms = ChatRoom.objects.select_related(
            'campaign',
            'client'
        ).all().order_by('-created_at')

        data = []

        for room in rooms:

       
            last_message = Message.objects.filter(
                room=room
            ).order_by('-timestamp').first()

            # Unread count
            unread_count = Message.objects.filter(
                room=room,
                is_read=False,
                sender_type='client'
            ).count()

            data.append({
                "room_id":       room.id,
                "campaign_id":   room.campaign.campaign_id,
                "campaign_name": room.campaign.campaign_name,
                "client_id":     room.client.client_id,
                "client_name":   room.client.name,
                "last_message":  last_message.content if last_message else None,
                "last_time":     last_message.timestamp.isoformat() if last_message else None,
                "unread_count":  unread_count,
            })

        return Response(data, status=200)

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def send_chat_file(request, campaign_id):

    try:

        # ── GET FILE ────────────────────────────────
        file        = request.FILES.get('file')
        sender_id   = request.data.get('sender_id')
        sender_type = request.data.get('sender_type')
        content     = request.data.get('content', '')

        if not file or not sender_id:
            return Response({"error": "file and sender_id required"}, status=400)

        # ── FIND ROOM ────────────────────────────────
        try:
            room = ChatRoom.objects.get(campaign__campaign_id=campaign_id)
        except ChatRoom.DoesNotExist:
            campaign = Campaign.objects.get(campaign_id=campaign_id)
            room = ChatRoom.objects.create(
                campaign=campaign,
                client=campaign.client
            )

        # ── DETECT FILE TYPE ─────────────────────────
        file_type = file.content_type  # e.g. image/jpeg, video/mp4

        if file_type.startswith('image/'):
            message_type = 'image'
        elif file_type.startswith('video/'):
            message_type = 'video'
        else:
            message_type = 'file'

        # ── FILE SIZE ─────────────────────────────────
        size = file.size
        if size < 1024 * 1024:
            file_size = f"{size // 1024} KB"
        else:
            file_size = f"{size / (1024 * 1024):.1f} MB"

        # ── SAVE MESSAGE ──────────────────────────────
        sender  = User.objects.get(id=sender_id)
        message = Message.objects.create(
            room=room,
            sender=sender,
            sender_type=sender_type,
            content=content,
            message_type=message_type,
            file=file,
            file_name=file.name,
            file_size=file_size,
        )

        # ── BROADCAST VIA WEBSOCKET ───────────────────
        channel_layer = get_channel_layer()
        room_group    = f"chat_{campaign_id}"

        file_url = request.build_absolute_uri(message.file.url)

        async_to_sync(channel_layer.group_send)(
            room_group,
            {
                'type':         'chat_message',
                'message_id':   message.id,
                'content':      content,
                'sender_id':    sender_id,
                'sender_type':  sender_type,
                'message_type': message_type,
                'file_url':     message.file.url,
                'file_name':    file.name,
                'file_size':    file_size,
                'timestamp':    message.timestamp.isoformat(),
            }
        )

        return Response({
            "message": "File sent successfully",
            "file_url": file_url,
            "message_type": message_type,
        }, status=201)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

#------------------------------New-------------------------------
@api_view(['PATCH'])
def update_creative_id(request):
    creative_type = request.data.get('type')      # 'standard' or 'third_party'
    creative_db_id = request.data.get('id')       # DB primary key (integer)
    creative_id_value = request.data.get('creative_id', '').strip()

    if not creative_db_id or not creative_type:
        return Response({"error": "id and type are required"}, status=400)

    try:
        if creative_type == 'standard':
            obj = Creative.objects.get(id=creative_db_id)
            obj.creative_id = creative_id_value
            obj.save()
            return Response({
                "message": "Creative ID Added Successfully",
                "id": obj.id,
                "creative_id": obj.creative_id,
            }, status=200)

        elif creative_type == 'third_party':
            obj = ThirdPartyCreative.objects.get(id=creative_db_id)
            obj.creative_id = creative_id_value
            obj.save()
            return Response({
                "message": "Creative ID Added Successfully",
                "id": obj.id,
                "creative_id": obj.creative_id,
            }, status=200)

        else:
            return Response({"error": "Invalid type. Use 'standard' or 'third_party'"}, status=400)

    except (Creative.DoesNotExist, ThirdPartyCreative.DoesNotExist):
        return Response({"error": "Creative not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
    

# ==============================
# ADD THESE TO YOUR views.py
# ==============================

import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import FileResponse, HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Campaign, CampaignExcel
from django.core.files.base import ContentFile


def build_campaign_excel(campaign, report_type='cpm', overrides=None) -> bytes:
    """
    overrides = { "LIBILL003": { "impressions": 50000, "start_date": "2026-06-01", ... } }
    """
    overrides = overrides or {}
    wb = Workbook()
    wb.remove(wb.active)

    # ── Styles (same as before) ──
    header_font  = Font(name='Arial', bold=True, size=10)
    value_font   = Font(name='Arial', size=10)
    label_fill   = PatternFill('solid', start_color='D9E1F2')
    title_fill   = PatternFill('solid', start_color='1F4E79')
    title_font   = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    border_side  = Side(style='thin', color='B8CCE4')
    thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    line_items = campaign.line_items.all()

    for li in line_items:
        sheet_name = str(li.line_item_id)[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 40

        ws.merge_cells('A1:B1')
        ws['A1'] = f"Campaign Report — {campaign.campaign_id}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 22

        insertion_order = getattr(campaign, 'insertion_order', None)
        is_cpc = report_type == 'cpc'
        has_creatives = (
            li.creatives_detail.exists() or li.third_party_creatives.exists()
        )
        creative_msg = (
            "Creatives will be shared by Creatives Team" if has_creatives else "Not yet received"
        )

        # ✅ Get overrides for this line item
        li_overrides = overrides.get(li.line_item_id, {})

        # ✅ Use override value if present, else use DB value
        def val(field, db_value):
            return li_overrides.get(field, db_value)

        impressions  = val('impressions', li.impressions)
        start_date   = val('start_date', str(li.start_date) if li.start_date else '')
        end_date     = val('end_date', str(li.end_date) if li.end_date else '')
        units        = val('units', li.units or '')
        ctr_raw      = val('ctr', li.ctr)
        viewability_raw = val('viewability', li.viewability)
        vcr_raw      = val('vcr', li.vcr)
        kpi_notes    = val('kpi_notes', li.kpi_notes or '')
        sitelist     = val('sitelist', '')

        ctr_display        = f"{ctr_raw}%" if ctr_raw else ''
        viewability_display = f"{viewability_raw}%" if viewability_raw else ''
        vcr_display        = f"{vcr_raw}%" if vcr_raw else ''

        rows = [
            ('Date of ID Setup',    campaign.created_at.strftime('%d-%B-%Y') if campaign.created_at else ''),
            ('Advertiser ID',       campaign.client.client_id if campaign.client else ''),
            ('Campaign ID',         f"{campaign.campaign_id}-A" if is_cpc else campaign.campaign_id),
            ('IO ID',               insertion_order.io_id if insertion_order else ''),
            ('IO Name',             campaign.campaign_name or ''),
            ('Clicks Booked' if is_cpc else 'Impressions Booked', impressions or ''),
            ('Start Date',          start_date),
            ('End Date',            end_date),
            ('Booked Budget' if is_cpc else 'Target CTR', ctr_display),
            ('Target CTR' if is_cpc else ctr_display, 'Booked Budget'),
            ('Line Item ID',        li.line_item_id or ''),
            ('Line Item Name',      li.line_item_name or ''),
            ('Ethnicity',           li.ethnicity or ''),
            ('Ad Format',           li.ad_format or ''),
            ('Geography',           _parse_geo(li.geo_targeting) if li.geo_targeting else ''),
            ('Market',              ''),
            ('Viewability',         viewability_display),
            ('Creative',            creative_msg),
            ('VCR',                 vcr_display),
            ('KPI',                 kpi_notes),
            ('Sitelist',            sitelist),
        ]

        for i, (label, value) in enumerate(rows, start=2):
            label_cell = ws.cell(row=i, column=1, value=label)
            value_cell = ws.cell(row=i, column=2, value=value)
            label_cell.font = header_font
            label_cell.fill = label_fill
            label_cell.alignment = left_align
            label_cell.border = thin_border
            value_cell.font = value_font
            value_cell.alignment = left_align
            value_cell.border = thin_border
            ws.row_dimensions[i].height = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

def _parse_geo(geo_targeting):
    """Parse geo_targeting field (JSON string or list) into readable string."""
    try:
        if isinstance(geo_targeting, str):
            data = json.loads(geo_targeting)
        else:
            data = geo_targeting
        parts = []
        for loc in (data if isinstance(data, list) else [data]):
            segment = ', '.join(filter(None, [
                loc.get('country', ''), loc.get('state', ''), loc.get('city', '')
            ]))
            if segment:
                parts.append(segment)
        return ' | '.join(parts) if parts else ''
    except Exception:
        return str(geo_targeting) if geo_targeting else ''


# ── Generate & store Excel ──────────────────────────────────────────────────

@api_view(['POST'])
def generate_campaign_excel(request, campaign_id):
    """
    Generate the Excel file for a campaign, save it to DB, return download URL.
    POST /generate_campaign_excel/<campaign_id>/
    """
    try:
        campaign = Campaign.objects.select_related(
            'client', 'insertion_order'
        ).prefetch_related(
            'line_items',
            'line_items__creatives_detail',          # ← add this
            'line_items__third_party_creatives',     # ← add this
        ).get(campaign_id=campaign_id)
    except Campaign.DoesNotExist:
        return Response({'error': 'Campaign not found'}, status=404)

    report_type = request.data.get('report_type', 'cpm')  # 'cpm' or 'cpc'
    excel_bytes = build_campaign_excel(campaign, report_type)
    filename    = f"{campaign_id}.xlsx"

    # Save / overwrite in DB
    excel_obj, _ = CampaignExcel.objects.get_or_create(campaign=campaign, report_type=report_type)
    excel_obj.excel_file.save(filename, ContentFile(excel_bytes), save=True)

    url = request.build_absolute_uri(excel_obj.excel_file.url)
    return Response({
        'message':      'Excel generated successfully',
        'campaign_id':  campaign_id,
        'download_url': url,
        'filename':     filename,
        'generated_at': excel_obj.generated_at.isoformat(),
    }, status=200)


@api_view(['GET'])
def get_campaigns_excel_list(request):
    campaigns = Campaign.objects.select_related('client').filter(
        approval_status='approved'
    ).prefetch_related('excel_reports', 'line_items').order_by('-created_at')

    data = []
    for c in campaigns:
        excel_map = {e.report_type: e for e in c.excel_reports.all()}

        # ✅ Check units across all line items
        all_units = [
            (li.units or '').upper()
            for li in c.line_items.all()
            if li.units
        ]

        has_cpm = any(u == 'CPM' for u in all_units)
        has_cpc = any(u == 'CPC' for u in all_units)

        # ✅ Decide which report types to generate
        if has_cpm and has_cpc:
            report_types = ['cpm', 'cpc']   # both units → both reports
        elif has_cpc:
            report_types = ['cpc']           # only CPC → one report (no -A suffix)
        else:
            report_types = ['cpm', 'cpc']   # only CPM or unknown → both reports

        for report_type in report_types:
            excel = excel_map.get(report_type)
            suffix = '-A' if report_type == 'cpc' else ''
            data.append({
                'campaign_id':      f"{c.campaign_id}{suffix}",
                'campaign_id_raw':  c.campaign_id,
                'report_type':      report_type,
                'campaign_name':    c.campaign_name,
                'client_name':      c.client.name if c.client else '',
                'client_id':        c.client.client_id if c.client else '',
                'start_date':       str(c.start_date) if c.start_date else '',
                'end_date':         str(c.end_date) if c.end_date else '',
                'line_items_count': c.line_items.count(),
                'excel_generated':  excel is not None,
                'excel_url':        request.build_absolute_uri(excel.excel_file.url) if excel else None,
                'generated_at':     excel.generated_at.isoformat() if excel else None,
                'publish_status':   excel.publish_status if excel else None,
            })

    return Response(data)
# ── Download saved Excel ────────────────────────────────────────────────────

@api_view(['GET'])
def download_campaign_excel(request, campaign_id):
    # ✅ Read report_type from query param
    report_type = request.query_params.get('report_type', 'cpm')

    try:
        excel_obj = CampaignExcel.objects.get(
            campaign__campaign_id=campaign_id,
            report_type=report_type   # ✅ Add this filter
        )
    except CampaignExcel.DoesNotExist:
        # Generate on-the-fly if not saved yet
        try:
            campaign = Campaign.objects.select_related(
                'client', 'insertion_order'
            ).prefetch_related(
                'line_items'
            ).get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({'error': 'Campaign not found'}, status=404)

        excel_bytes = build_campaign_excel(campaign, report_type)
        response = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{campaign_id}_{report_type}.xlsx"'
        return response

    response = FileResponse(
        excel_obj.excel_file.open('rb'),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{campaign_id}_{report_type}.xlsx"'
    return response

@api_view(['POST'])
def save_excel_edits_to_db(request, campaign_id):
    try:
        report_type  = request.data.get('report_type', 'cpm')
        line_item_id = request.data.get('line_item_id')

        if not line_item_id:
            return Response({'error': 'line_item_id is required'}, status=400)

        # ── 1. Get CampaignExcel record ──────────────────────────────────
        try:
            excel_obj = CampaignExcel.objects.get(
                campaign__campaign_id=campaign_id,
                report_type=report_type
            )
        except CampaignExcel.DoesNotExist:
            return Response({'error': 'Excel record not found. Generate first.'}, status=404)

        # ── 2. Get LineItem ───────────────────────────────────────────────
        try:
            line_item = LineItem.objects.get(line_item_id=line_item_id)
        except LineItem.DoesNotExist:
            return Response({'error': 'LineItem not found'}, status=404)

        # ── 3. Get Campaign ───────────────────────────────────────────────
        try:
            campaign = Campaign.objects.select_related(
                'client', 'insertion_order'
            ).prefetch_related(
                'line_items',
                'line_items__creatives_detail',
                'line_items__third_party_creatives',
            ).get(campaign_id=campaign_id)
        except Campaign.DoesNotExist:
            return Response({'error': 'Campaign not found'}, status=404)

        editable_fields = [
            'impressions', 'start_date', 'end_date', 'units',
            'ctr', 'viewability', 'vcr', 'kpi_notes', 'sitelist'
        ]

        # ── 4. Build override dict ────────────────────────────────────────
        li_override = excel_obj.line_item_overrides.get(line_item_id, {})
        for field in editable_fields:
            if field in request.data and request.data[field] is not None:
                li_override[field] = request.data[field]

        # ── 5. Save overrides to CampaignExcel.line_item_overrides ────────
        overrides = excel_obj.line_item_overrides.copy()
        overrides[line_item_id] = li_override
        excel_obj.line_item_overrides = overrides
        excel_obj.save(update_fields=['line_item_overrides'])

        # ── 6. Update LineItem table (except sitelist) ────────────────────
        if 'impressions' in li_override:
            try:
                line_item.impressions = int(
                    str(li_override['impressions']).replace(',', '').strip()
                )
            except (ValueError, TypeError):
                pass

        if 'start_date' in li_override and li_override['start_date']:
            line_item.start_date = parse_date(str(li_override['start_date']))

        if 'end_date' in li_override and li_override['end_date']:
            line_item.end_date = parse_date(str(li_override['end_date']))

        if 'units' in li_override and li_override['units']:
            line_item.units = str(li_override['units']).strip()

        if 'ctr' in li_override:
            try:
                line_item.ctr = float(
                    str(li_override['ctr']).replace('%', '').strip()
                )
            except (ValueError, TypeError):
                pass

        if 'viewability' in li_override:
            try:
                line_item.viewability = float(
                    str(li_override['viewability']).replace('%', '').strip()
                )
            except (ValueError, TypeError):
                pass

        if 'vcr' in li_override:
            try:
                line_item.vcr = float(
                    str(li_override['vcr']).replace('%', '').strip()
                )
            except (ValueError, TypeError):
                pass

        if 'kpi_notes' in li_override:
            line_item.kpi_notes = str(li_override['kpi_notes'])

        line_item.save()

        # ── 7. Save to CampaignLineItemExcel table ────────────────────────
        from .models import CampaignLineItemExcel

        excel_li_obj, _ = CampaignLineItemExcel.objects.update_or_create(
            line_item=line_item,
            report_type=report_type,
            defaults={
                'campaign': campaign,
                'impressions': line_item.impressions,
                'start_date': line_item.start_date,
                'end_date': line_item.end_date,
                'units': line_item.units,
                'ctr': line_item.ctr,
                'viewability': line_item.viewability,
                'vcr': line_item.vcr,
                'kpi_notes': line_item.kpi_notes,
                'sitelist': li_override.get('sitelist', ''),
            }
        )

        # ── 8. Re-generate Excel with updated overrides ───────────────────
        excel_bytes = build_campaign_excel(campaign, report_type, overrides=overrides)
        filename = f"{campaign_id}_{report_type}.xlsx"
        excel_obj.excel_file.save(filename, ContentFile(excel_bytes), save=True)

        url = request.build_absolute_uri(excel_obj.excel_file.url)
        return Response({
            'message': 'Saved to LineItem, CampaignLineItemExcel & Excel regenerated',
            'download_url': url,
            'line_item_overrides': overrides,
            'excel_line_item_id': excel_li_obj.id,
        }, status=200)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def publish_campaign_excel(request, campaign_id):
    report_type = request.data.get('report_type', 'cpm')
    try:
        excel_obj = CampaignExcel.objects.get(
            campaign__campaign_id=campaign_id,
            report_type=report_type
        )
    except CampaignExcel.DoesNotExist:
        return Response({'error': 'Excel not found. Generate first.'}, status=404)

    excel_obj.publish_status = 'published'
    excel_obj.save(update_fields=['publish_status'])

    return Response({
        'message': 'Excel published successfully',
        'campaign_id': campaign_id,
        'report_type': report_type,
        'publish_status': excel_obj.publish_status,
    }, status=200)
    
    


from weasyprint import HTML as WeasyHTML
from django.core.files.base import ContentFile
from .pdf_templates import build_io_html, build_invoice_html
from .models import InsertionOrder
from .models import Invoice

# ── Generate IO PDF ──────────────────────────────────────────────────────────
@api_view(['POST'])
def generate_io_pdf(request, campaign_id):
    try:
        # CORRECT — use select_related for OneToOne
        campaign = Campaign.objects.select_related(
            'client', 'insertion_order'
        ).prefetch_related(
            'line_items__creatives_detail',
            'line_items__third_party_creatives',
        ).get(campaign_id=campaign_id)
    except Campaign.DoesNotExist:
        return Response({'error': 'Campaign not found'}, status=404)

    # Get or create IO record
    io_obj, _ = InsertionOrder.objects.get_or_create(
        campaign=campaign,
        defaults={'client': campaign.client}
    )

    # Fetch client details
    try:
        client = Client.objects.select_related(
            'billing', 'ownership'
        ).prefetch_related('contacts').get(pk=campaign.client.pk)
    except Client.DoesNotExist:
        client = None

    # Build HTML and convert to PDF
    html_string = build_io_html(campaign, client, io_obj.io_id)
    pdf_bytes = WeasyHTML(string=html_string).write_pdf()

    # Save PDF to model
    filename = f"{io_obj.io_id}_{campaign_id}.pdf"
    io_obj.pdf_file.save(filename, ContentFile(pdf_bytes), save=True)

    url = request.build_absolute_uri(io_obj.pdf_file.url)
    return Response({
        'message': 'IO PDF generated successfully',
        'io_id': io_obj.io_id,
        'download_url': url,
    }, status=200)


# ── Generate Invoice PDF ─────────────────────────────────────────────────────
@api_view(['POST'])
def generate_invoice_pdf(request, campaign_id):
    try:
        campaign = Campaign.objects.select_related(
            'client', 'invoice'
        ).prefetch_related(
            'line_items__creatives_detail',
            'line_items__third_party_creatives',
        ).get(campaign_id=campaign_id)
    except Campaign.DoesNotExist:
        return Response({'error': 'Campaign not found'}, status=404)

    # ── Check approval ──
    if campaign.approval_status != 'approved':
        return Response({'error': 'Campaign is not approved'}, status=400)

    # ── Check end date has passed ──
    from datetime import date
    if not campaign.end_date:
        return Response({'error': 'Campaign has no end date'}, status=400)

    if campaign.end_date > date.today():
        return Response({
            'error': 'Invoice cannot be generated before campaign end date',
            'end_date': str(campaign.end_date)
        }, status=400)

    # ── Get or create invoice record ──
    invoice_obj, _ = Invoice.objects.get_or_create(
        campaign=campaign,
        defaults={'client': campaign.client}
    )

    # ── Fetch client details ──
    try:
        client = Client.objects.select_related(
            'billing', 'ownership'
        ).prefetch_related('contacts').get(pk=campaign.client.pk)
    except Client.DoesNotExist:
        client = None

    # ── Build HTML and convert to PDF ──
    html_string = build_invoice_html(campaign, client)
    pdf_bytes = WeasyHTML(string=html_string).write_pdf()

    # ── Save PDF to model ──
    filename = f"{invoice_obj.invoice_id}_{campaign_id}.pdf"
    invoice_obj.pdf_file.save(filename, ContentFile(pdf_bytes), save=True)

    url = request.build_absolute_uri(invoice_obj.pdf_file.url)
    return Response({
        'message': 'Invoice PDF generated successfully',
        'invoice_id': invoice_obj.invoice_id,
        'download_url': url,
    }, status=200)

# ── Download saved IO PDF ────────────────────────────────────────────────────
@api_view(['GET'])
def download_io_pdf(request, campaign_id):
    try:
        io_obj = InsertionOrder.objects.get(campaign__campaign_id=campaign_id)
    except InsertionOrder.DoesNotExist:
        return Response({'error': 'IO not found'}, status=404)

    if not io_obj.pdf_file:
        return Response({'error': 'PDF not generated yet'}, status=404)

    response = FileResponse(
        io_obj.pdf_file.open('rb'),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="{io_obj.io_id}.pdf"'
    return response


# ── Download saved Invoice PDF ───────────────────────────────────────────────
@api_view(['GET'])
def download_invoice_pdf(request, campaign_id):
    try:
        invoice_obj = Invoice.objects.get(campaign__campaign_id=campaign_id)
    except Invoice.DoesNotExist:
        return Response({'error': 'Invoice not found'}, status=404)

    if not invoice_obj.pdf_file:
        return Response({'error': 'PDF not generated yet'}, status=404)

    response = FileResponse(
        invoice_obj.pdf_file.open('rb'),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="{invoice_obj.invoice_id}.pdf"'
    return response

# ── ADD THIS to your views.py ──────────────────────────────────────────────────

@api_view(['GET'])
def get_io_list_by_client(request, client_id):
    """
    Returns all approved campaigns for a client with their IO/PDF status.
    GET /get_io_list_by_client/<client_id>/
    """
    campaigns = Campaign.objects.select_related(
        'client', 'insertion_order'
    ).filter(
        client__client_id=client_id,
        approval_status='approved'
    ).prefetch_related('line_items').order_by('-created_at')

    data = []
    for c in campaigns:
        io = getattr(c, 'insertion_order', None)
        data.append({
            'campaign_id':      c.campaign_id,
            'campaign_name':    c.campaign_name,
            'advertiser':       c.advertiser or '',
            'client_name':      c.client.name if c.client else '',
            'client_id':        c.client.client_id if c.client else '',
            'start_date':       str(c.start_date) if c.start_date else '',
            'end_date':         str(c.end_date) if c.end_date else '',
            'campaign_type':    c.campaign_type or '',
            'line_items_count': c.line_items.count(),
            'io_id':            io.io_id if io else None,
            # pdf_generated = IO exists AND has a pdf_file saved
            'pdf_generated':    bool(io and io.pdf_file),
            'pdf_url':          request.build_absolute_uri(io.pdf_file.url) if io and io.pdf_file else None,
            'created_at':       c.created_at.isoformat() if c.created_at else '',
        })

    return Response(data)



# ── ADD THIS to your views.py ──────────────────────────────────────────────────

@api_view(['GET'])
def get_invoice_list_by_client(request, client_id):
    """
    Returns all approved campaigns for a client where end_date has passed,
    with their Invoice/PDF status.
    GET /get_invoice_list_by_client/<client_id>/
    """
    from datetime import date as date_today

    campaigns = Campaign.objects.select_related(
        'client', 'invoice'
    ).filter(
        client__client_id=client_id,
        approval_status='approved',
        end_date__lte=date_today.today(),   # only campaigns whose end date has passed
    ).prefetch_related('line_items').order_by('-created_at')

    data = []
    for c in campaigns:
        inv = getattr(c, 'invoice', None)
        data.append({
            'campaign_id':      c.campaign_id,
            'campaign_name':    c.campaign_name,
            'advertiser':       c.advertiser or '',
            'client_name':      c.client.name if c.client else '',
            'client_id':        c.client.client_id if c.client else '',
            'start_date':       str(c.start_date) if c.start_date else '',
            'end_date':         str(c.end_date) if c.end_date else '',
            'campaign_type':    c.campaign_type or '',
            'line_items_count': c.line_items.count(),
            'invoice_id':       inv.invoice_id if inv else None,
            # pdf_generated = Invoice exists AND has a pdf_file saved
            'pdf_generated':    bool(inv and inv.pdf_file),
            'pdf_url':          request.build_absolute_uri(inv.pdf_file.url) if inv and inv.pdf_file else None,
            'created_at':       c.created_at.isoformat() if c.created_at else '',
        })

    return Response(data)

